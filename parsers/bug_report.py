"""
Парсер Worqen_QA_Full_Report_.xlsx з tier-системою.
"""

import io
import re
from typing import Optional
from openpyxl import load_workbook

from drive_client import download_file
from config import FILE_IDS, SCAN_BUG_DESC_TRUNCATE


BUG_SHEET = "Bug"
FICHA_SHEET = "Ficha"

BUG_COLUMNS = [
    "ID", "Тип", "Пріоритет", "Джерело", "Статус",
    "Опис", "Де", "Очікувана поведінка",
    "Рекомендація для Романа", "Посилання на скрін",
    "Зафіксоване рішення", "Примітки",
]

BUG_SCAN_FIELDS = ["ID", "Тип", "Пріоритет", "Статус", "Опис"]
BUG_AUDIT_FIELDS = BUG_SCAN_FIELDS + [
    "Де", "Очікувана поведінка", "Рекомендація для Романа",
    "Зафіксоване рішення", "Примітки",
]

FICHA_COLUMNS = [
    "#", "Назва", "Опис", "Пріоритет",
    "Перенесено в US", "Посилання", "Нотатки", "Дата",
]

FICHA_SCAN_FIELDS = ["#", "Назва", "Пріоритет", "Перенесено в US"]
FICHA_AUDIT_FIELDS = FICHA_SCAN_FIELDS + ["Опис", "Нотатки", "Дата"]


def _load_sheet_with_header_row2(sheet_name, columns):
    content = download_file(FILE_IDS["qa_report"])
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    header_row = rows[1]
    col_idx = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip()
        for expected in columns:
            if cell_str.startswith(expected):
                if expected not in col_idx:
                    col_idx[expected] = i
                break

    items = []
    for row in rows[2:]:
        if all(cell is None for cell in row):
            continue
        item = {}
        for col_name in columns:
            if col_name not in col_idx:
                item[col_name] = None
                continue
            idx = col_idx[col_name]
            value = row[idx] if idx < len(row) else None
            item[col_name] = None if value is None else str(value)
        primary = item.get("ID") or item.get("#")
        if not primary:
            continue
        items.append(item)
    return items


def _truncate(value, limit):
    if value is None:
        return None
    if len(value) <= limit:
        return value
    return value[: limit - 3] + "..."


def _apply_bug_tier(bug, tier):
    if tier == "full":
        return bug
    if tier == "audit":
        return {k: bug.get(k) for k in BUG_AUDIT_FIELDS if k in bug}
    result = {k: bug.get(k) for k in BUG_SCAN_FIELDS if k in bug}
    if "Опис" in result and result["Опис"]:
        result["Опис"] = _truncate(result["Опис"], SCAN_BUG_DESC_TRUNCATE)
    return result


def _apply_ficha_tier(ficha, tier):
    if tier == "full":
        return ficha
    if tier == "audit":
        return {k: ficha.get(k) for k in FICHA_AUDIT_FIELDS if k in ficha}
    return {k: ficha.get(k) for k in FICHA_SCAN_FIELDS if k in ficha}


def list_bugs(priority=None, status=None, bug_type=None, search=None, tier="scan"):
    bugs = _load_sheet_with_header_row2(BUG_SHEET, BUG_COLUMNS)
    result = []
    for b in bugs:
        if priority and (b.get("Пріоритет") or "").upper() != priority.upper():
            continue
        if status and (b.get("Статус") or "").lower() != status.lower():
            continue
        if bug_type and (b.get("Тип") or "").lower() != bug_type.lower():
            continue
        if search:
            haystack = " ".join(
                str(b.get(k) or "")
                for k in ("Опис", "Де", "Очікувана поведінка",
                         "Рекомендація для Романа", "Примітки")
            ).lower()
            if search.lower() not in haystack:
                continue
        result.append(_apply_bug_tier(b, tier))
    return result


def get_bug(bug_id):
    bid = bug_id.strip().upper()
    for b in _load_sheet_with_header_row2(BUG_SHEET, BUG_COLUMNS):
        if (b.get("ID") or "").strip().upper() == bid:
            return b
    return None


def get_next_bug_id():
    max_num = 0
    pattern = re.compile(r"^BUG-(\d+)$", re.IGNORECASE)
    for b in _load_sheet_with_header_row2(BUG_SHEET, BUG_COLUMNS):
        m = pattern.match((b.get("ID") or "").strip())
        if m:
            num = int(m.group(1))
            if num > max_num:
                max_num = num
    return f"BUG-{max_num + 1:03d}"


def bugs_summary(level="full"):
    bugs = _load_sheet_with_header_row2(BUG_SHEET, BUG_COLUMNS)
    summary = {
        "total": len(bugs),
        "by_priority": {},
        "by_status": {},
        "by_type": {},
        "by_priority_status": {},
    }
    for b in bugs:
        pr = b.get("Пріоритет") or "(no priority)"
        st = b.get("Статус") or "(no status)"
        tp = b.get("Тип") or "(no type)"
        summary["by_priority"][pr] = summary["by_priority"].get(pr, 0) + 1
        summary["by_status"][st] = summary["by_status"].get(st, 0) + 1
        summary["by_type"][tp] = summary["by_type"].get(tp, 0) + 1
        if pr not in summary["by_priority_status"]:
            summary["by_priority_status"][pr] = {}
        summary["by_priority_status"][pr][st] = (
            summary["by_priority_status"][pr].get(st, 0) + 1
        )
    if level == "minimal":
        return {"total": summary["total"], "by_priority": summary["by_priority"]}
    return summary


def list_ficha(priority=None, transferred=None, search=None, tier="scan"):
    items = _load_sheet_with_header_row2(FICHA_SHEET, FICHA_COLUMNS)
    result = []
    for f in items:
        if priority and (f.get("Пріоритет") or "").lower() != priority.lower():
            continue
        if transferred is True and not f.get("Перенесено в US"):
            continue
        if transferred is False and f.get("Перенесено в US"):
            continue
        if search:
            haystack = " ".join(
                str(f.get(k) or "") for k in ("Назва", "Опис", "Нотатки")
            ).lower()
            if search.lower() not in haystack:
                continue
        result.append(_apply_ficha_tier(f, tier))
    return result


def get_ficha(ficha_id):
    norm = ficha_id.strip().lower()
    for f in _load_sheet_with_header_row2(FICHA_SHEET, FICHA_COLUMNS):
        if (f.get("#") or "").strip().lower() == norm:
            return f
    return None


def get_next_ficha_id():
    max_num = 0
    pattern = re.compile(r"^Ficha-(\d+)$", re.IGNORECASE)
    for f in _load_sheet_with_header_row2(FICHA_SHEET, FICHA_COLUMNS):
        m = pattern.match((f.get("#") or "").strip())
        if m:
            num = int(m.group(1))
            if num > max_num:
                max_num = num
    return f"Ficha-{max_num + 1:03d}"


def get_raw_bugs():
    return _load_sheet_with_header_row2(BUG_SHEET, BUG_COLUMNS)


def get_raw_ficha():
    return _load_sheet_with_header_row2(FICHA_SHEET, FICHA_COLUMNS)


def get_bug_row_index_by_id(bug_id: str) -> Optional[int]:
    """
    Повертає 1-based номер рядка в xlsx для вказаного BUG-ID.
    Header у row 2, дані з row 3.
    None якщо не знайдено.
    """
    bid = bug_id.strip().upper()
    content = download_file(FILE_IDS["qa_report"])
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if BUG_SHEET not in wb.sheetnames:
        return None
    ws = wb[BUG_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return None
    header_row = rows[1]
    id_col = None
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        if str(cell).strip().startswith("ID"):
            id_col = i
            break
    if id_col is None:
        return None
    for i, row in enumerate(rows[2:], start=3):
        if id_col >= len(row):
            continue
        cell = row[id_col]
        if cell is None:
            continue
        if str(cell).strip().upper() == bid:
            return i
    return None


def get_last_bug_row_index() -> int:
    """
    Повертає 1-based номер ОСТАННЬОГО рядка з даними у Bug sheet.
    Якщо даних немає — повертає 2 (header row).
    """
    content = download_file(FILE_IDS["qa_report"])
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[BUG_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    last = 2
    for i, row in enumerate(rows[2:], start=3):
        if all(c is None for c in row):
            continue
        last = i
    return last


def find_missing_bug_ids() -> list[str]:
    """
    Повертає список пропусків у послідовності BUG-ID.
    Якщо пропусків немає — порожній список.
    """
    pattern = re.compile(r"^BUG-(\d+)$", re.IGNORECASE)
    existing_nums = set()
    max_num = 0
    for b in _load_sheet_with_header_row2(BUG_SHEET, BUG_COLUMNS):
        m = pattern.match((b.get("ID") or "").strip())
        if m:
            num = int(m.group(1))
            existing_nums.add(num)
            if num > max_num:
                max_num = num
    if max_num == 0:
        return []
    missing = []
    for n in range(1, max_num + 1):
        if n not in existing_nums:
            missing.append(f"BUG-{n:03d}")
    return missing


def is_bug_id_free(bug_id: str) -> tuple[bool, Optional[str]]:
    """
    Чи вільний вказаний BUG-ID.
    Повертає (is_free, occupant_desc_preview).
    """
    bid = bug_id.strip().upper()
    for b in _load_sheet_with_header_row2(BUG_SHEET, BUG_COLUMNS):
        if (b.get("ID") or "").strip().upper() == bid:
            opis = b.get("Опис") or ""
            preview = opis[:100] + ("..." if len(opis) > 100 else "")
            return False, preview
    return True, None


_BUG_ID_PATTERN = re.compile(r"^BUG-\d{3,4}$", re.IGNORECASE)


def is_valid_bug_id_format(bug_id: str) -> bool:
    """Перевірка формату: BUG-NNN або BUG-NNNN."""
    if not bug_id:
        return False
    return bool(_BUG_ID_PATTERN.match(bug_id.strip()))
