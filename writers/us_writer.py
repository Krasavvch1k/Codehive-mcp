"""
Writer для User Stories xlsx.

Дві операції:
- update_story(story_id, fields)  — точкова заміна whitelist-полів
- create_story(fields)            — додає новий рядок з next_us_id

Безпека:
- drive sync check перед записом
- auto snapshot якщо сьогодні ще не було
- whitelist полів (Status/Priority/Version/Est. day/Epic/ID — не чіпає при update)
- стиль нового рядка копіюється з останнього рядка з даними

Після успіху — clear_cache для user_stories file_id.
"""

import io
from copy import copy
from typing import Optional

from openpyxl import load_workbook

from drive_client import (
    download_file,
    upload_file_content,
    fetch_current_drive_modified,
    clear_cache,
)
from projects.worqen.config import (
    FILE_IDS,
    XLSX_MIME,
    US_EDITABLE_FIELDS,
    US_REQUIRED_ON_CREATE,
    US_ALL_FIELDS,
)
from parsers.user_stories import (
    SHEET_NAME,
    get_next_us_id,
    get_row_index_by_id,
    is_us_id_free,
    is_valid_us_id_format,
)
from writers.safety import (
    SafetyError,
    check_drive_unchanged,
    ensure_today_snapshot,
)


def _validate_update_fields(fields: dict) -> None:
    """Кидає SafetyError якщо у fields є неприпустимі ключі."""
    if not fields:
        raise SafetyError(
            "fields порожній — нічого редагувати.",
            kind="empty_fields",
        )
    bad = [k for k in fields.keys() if k not in US_EDITABLE_FIELDS]
    if bad:
        raise SafetyError(
            f"Поля {bad} не дозволено редагувати через MCP. "
            f"Дозволені: {US_EDITABLE_FIELDS}. "
            f"Status / Priority / Version / Est. day / Epic — Микита редагує руками.",
            kind="forbidden_field",
        )


def _validate_create_fields(fields: dict) -> None:
    """Перевірка створення: всі required поля присутні + усі ключі валідні."""
    if not fields:
        raise SafetyError("fields порожній.", kind="empty_fields")

    missing = [k for k in US_REQUIRED_ON_CREATE if not fields.get(k)]
    if missing:
        raise SafetyError(
            f"Бракує обов'язкових полів при створенні US: {missing}. "
            f"Required: {US_REQUIRED_ON_CREATE}.",
            kind="missing_required",
        )

    # ID забороняємо передавати у fields — для конкретного ID є параметр use_id
    if "ID" in fields:
        raise SafetyError(
            "ID не передається у fields. Для конкретного ID використовуй параметр use_id; "
            "інакше ID присвоюється автоматично через next_us_id.",
            kind="forbidden_field",
        )

    bad = [k for k in fields.keys() if k not in US_ALL_FIELDS]
    if bad:
        raise SafetyError(
            f"Невідомі поля: {bad}. Дозволені: {US_ALL_FIELDS}.",
            kind="unknown_field",
        )


def _open_workbook_writable() -> tuple[bytes, "Workbook", "Worksheet", dict[str, int]]:
    """
    Завантажує xlsx з Drive, відкриває openpyxl у write-mode,
    повертає (raw_bytes, wb, ws, header_map).

    header_map: {column_name: 1-based column index}
    Для US headers у row 1.
    """
    file_id = FILE_IDS["user_stories"]
    raw = download_file(file_id, force_refresh=True)
    wb = load_workbook(io.BytesIO(raw))
    if SHEET_NAME not in wb.sheetnames:
        raise SafetyError(
            f"Sheet '{SHEET_NAME}' не знайдено у user_stories.xlsx.",
            kind="sheet_missing",
        )
    ws = wb[SHEET_NAME]

    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if name and name not in header_map:
            header_map[name] = col_idx

    return raw, wb, ws, header_map


def _save_and_upload(wb) -> dict:
    """Серіалізує workbook у байти і заливає назад у Drive."""
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return upload_file_content(
        FILE_IDS["user_stories"],
        out.getvalue(),
        XLSX_MIME,
    )


def update_story(story_id: str, fields: dict, force_overwrite: bool = False) -> dict:
    """
    Оновлює whitelist-поля існуючої US.
    Перевіряє drive sync, робить snapshot, пише, інвалідує кеш.
    """
    _validate_update_fields(fields)

    snapshot_info = ensure_today_snapshot()

    file_id = FILE_IDS["user_stories"]
    baseline_modified = fetch_current_drive_modified(file_id)

    # Резолвимо рядок по ID на свіжій копії
    row_idx = get_row_index_by_id(story_id)
    if row_idx is None:
        raise SafetyError(
            f"US '{story_id}' не знайдено у файлі.",
            kind="not_found",
        )

    # Drive sync check ПЕРЕД записом (між нашим read і моментом write
    # файл міг змінитись). При force_overwrite=True — пропускаємо.
    if not force_overwrite:
        check_drive_unchanged(file_id, baseline_modified)

    _raw, wb, ws, header_map = _open_workbook_writable()

    # Повторно валідуємо що рядок все ще цей самий ID (паранойя)
    id_col = header_map.get("ID")
    if id_col is None:
        raise SafetyError("Колонка 'ID' не знайдена у header.", kind="header_missing")
    actual_id = ws.cell(row=row_idx, column=id_col).value
    if actual_id is None or str(actual_id).strip().upper() != story_id.strip().upper():
        raise SafetyError(
            f"Рядок {row_idx} більше не містить {story_id} (там '{actual_id}'). "
            f"Файл змінився між read і write — повтори операцію.",
            kind="row_mismatch",
        )

    updated_fields = []
    for field, value in fields.items():
        col_idx = header_map.get(field)
        if col_idx is None:
            raise SafetyError(
                f"Колонка '{field}' не знайдена у header.",
                kind="header_missing",
            )
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        # Багаторядкові поля → wrap_text
        if isinstance(value, str) and "\n" in value:
            from openpyxl.styles import Alignment
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical=existing.vertical or "top",
                wrap_text=True,
            )
        updated_fields.append(field)

    upload_meta = _save_and_upload(wb)
    clear_cache(file_id)

    return {
        "id": story_id,
        "row": row_idx,
        "updated_fields": updated_fields,
        "snapshot": snapshot_info,
        "drive_modified_after": upload_meta.get("modifiedTime"),
        "force_overwrite_used": force_overwrite,
    }


def create_story(fields: dict, use_id: str | None = None, force_overwrite: bool = False) -> dict:
    """
    Створює нову US.
    Якщо use_id переданий — використовує цей ID (після валідації і перевірки вільності),
    інакше — присвоює через next_us_id.
    У будь-якому випадку запис йде в КІНЕЦЬ файлу (не у середину пропусків).
    Стиль нового рядка копіюється з останнього рядка з даними.
    """
    _validate_create_fields(fields)

    # Швидка валідація формату use_id ДО будь-яких Drive-операцій
    if use_id is not None:
        use_id = use_id.strip().upper()
        if not is_valid_us_id_format(use_id):
            raise SafetyError(
                f"Невалідний формат use_id='{use_id}'. Очікую US-NNN або US-NNNN.",
                kind="invalid_id_format",
            )

    snapshot_info = ensure_today_snapshot()

    file_id = FILE_IDS["user_stories"]
    baseline_modified = fetch_current_drive_modified(file_id)

    if use_id is None:
        new_id = get_next_us_id()
    else:
        new_id = use_id

    if not force_overwrite:
        check_drive_unchanged(file_id, baseline_modified)

    _raw, wb, ws, header_map = _open_workbook_writable()

    id_col = header_map.get("ID")
    if id_col is None:
        raise SafetyError("Колонка 'ID' не знайдена.", kind="header_missing")

    # Перевірка вільності ID на вже завантаженій книзі.
    # Для use_id — повідомлення з occupant Title; для auto — race-condition.
    new_id_upper = new_id.strip().upper()
    title_col = header_map.get("Title")
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[id_col - 1]
        if cell.value is None:
            continue
        if str(cell.value).strip().upper() == new_id_upper:
            occupant_title = None
            if title_col is not None:
                occupant_title = ws.cell(row=cell.row, column=title_col).value
            if use_id is not None:
                raise SafetyError(
                    f"ID {new_id} вже зайнятий (Title: '{occupant_title}'). "
                    f"Або вибери інший вільний ID, або викликай create без use_id.",
                    kind="id_collision",
                )
            raise SafetyError(
                f"ID {new_id} зайнятий — race condition (Title: '{occupant_title}'). "
                f"Повтори операцію.",
                kind="id_collision",
            )

    # Знаходимо номер останнього рядка з даними (для стилю + щоб писати ПІСЛЯ нього)
    last_row = 1
    for row in ws.iter_rows(min_row=2, values_only=False):
        if all(c.value is None for c in row):
            continue
        last_row = row[0].row
    new_row = last_row + 1 if last_row >= 2 else 2

    # Копіюємо стиль з last_row у new_row (по всіх колонках)
    if last_row >= 2:
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=last_row, column=col_idx)
            dst = ws.cell(row=new_row, column=col_idx)
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

    # Заповнюємо: ID + всі передані поля
    full_fields = {"ID": new_id, **fields}
    written_fields = []
    for field, value in full_fields.items():
        col_idx = header_map.get(field)
        if col_idx is None:
            raise SafetyError(
                f"Колонка '{field}' не знайдена у header.",
                kind="header_missing",
            )
        cell = ws.cell(row=new_row, column=col_idx)
        cell.value = value
        if isinstance(value, str) and "\n" in value:
            from openpyxl.styles import Alignment
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical=existing.vertical or "top",
                wrap_text=True,
            )
        written_fields.append(field)

    upload_meta = _save_and_upload(wb)
    clear_cache(file_id)

    return {
        "id": new_id,
        "row": new_row,
        "written_fields": written_fields,
        "snapshot": snapshot_info,
        "drive_modified_after": upload_meta.get("modifiedTime"),
        "force_overwrite_used": force_overwrite,
    }
