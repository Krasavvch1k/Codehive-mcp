"""
Writer для Bug Report xlsx.

Дві операції:
- update_bug(bug_id, fields)  — точкова заміна whitelist-полів
- create_bug(fields)          — додає новий рядок з next_bug_id

Безпека:
- drive sync check перед записом
- auto snapshot якщо сьогодні ще не було
- whitelist полів (Тип/Пріоритет/Статус/Джерело/ID — не чіпає при update)
- стиль нового рядка копіюється з останнього рядка з даними

Особливість vs us_writer: header у row 2 (не row 1), дані з row 3.
"""

import io
from copy import copy

from openpyxl import load_workbook

from drive_client import (
    download_file,
    upload_file_content,
    fetch_current_drive_modified,
    clear_cache,
)
from config import (
    FILE_IDS,
    XLSX_MIME,
    BUG_EDITABLE_FIELDS,
    BUG_REQUIRED_ON_CREATE,
    BUG_ALL_FIELDS,
)
from parsers.bug_report import (
    BUG_SHEET,
    BUG_COLUMNS,
    get_next_bug_id,
    get_bug_row_index_by_id,
    is_bug_id_free,
    is_valid_bug_id_format,
)
from writers.safety import (
    SafetyError,
    check_drive_unchanged,
    ensure_today_snapshot,
)
from writers.writes_log import log_write


HEADER_ROW = 2
DATA_START_ROW = 3


def _validate_update_fields(fields: dict) -> None:
    if not fields:
        raise SafetyError("fields порожній.", kind="empty_fields")
    bad = [k for k in fields.keys() if k not in BUG_EDITABLE_FIELDS]
    if bad:
        raise SafetyError(
            f"Поля {bad} не дозволено редагувати через MCP. "
            f"Дозволені: {BUG_EDITABLE_FIELDS}. "
            f"Тип / Пріоритет / Статус / Джерело — Микита редагує руками.",
            kind="forbidden_field",
        )


def _validate_create_fields(fields: dict) -> None:
    if not fields:
        raise SafetyError("fields порожній.", kind="empty_fields")

    missing = [k for k in BUG_REQUIRED_ON_CREATE if not fields.get(k)]
    if missing:
        raise SafetyError(
            f"Бракує обов'язкових полів при створенні BUG: {missing}. "
            f"Required: {BUG_REQUIRED_ON_CREATE}.",
            kind="missing_required",
        )

    if "ID" in fields:
        raise SafetyError(
            "ID не передається у fields. Для конкретного ID використовуй параметр use_id; "
            "інакше ID присвоюється автоматично через next_bug_id.",
            kind="forbidden_field",
        )

    bad = [k for k in fields.keys() if k not in BUG_ALL_FIELDS]
    if bad:
        raise SafetyError(
            f"Невідомі поля: {bad}. Дозволені: {BUG_ALL_FIELDS}.",
            kind="unknown_field",
        )


def _open_workbook_writable():
    """
    Завантажує qa_report з Drive, відкриває openpyxl у write-mode,
    повертає (raw_bytes, wb, ws, header_map).

    header_map: {column_name: 1-based column index}.
    Header у row 2. Маппінг — по startswith (як у парсері).
    """
    file_id = FILE_IDS["qa_report"]
    raw = download_file(file_id, force_refresh=True)
    wb = load_workbook(io.BytesIO(raw))
    if BUG_SHEET not in wb.sheetnames:
        raise SafetyError(
            f"Sheet '{BUG_SHEET}' не знайдено у qa_report.xlsx.",
            kind="sheet_missing",
        )
    ws = wb[BUG_SHEET]

    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[HEADER_ROW], start=1):
        if cell.value is None:
            continue
        cell_str = str(cell.value).strip()
        for expected in BUG_COLUMNS:
            if cell_str.startswith(expected) and expected not in header_map:
                header_map[expected] = col_idx
                break

    return raw, wb, ws, header_map


def _save_and_upload(wb) -> dict:
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return upload_file_content(
        FILE_IDS["qa_report"],
        out.getvalue(),
        XLSX_MIME,
    )


def update_bug(bug_id: str, fields: dict, force_overwrite: bool = False) -> dict:
    """Оновлює whitelist-поля існуючого BUG."""
    _validate_update_fields(fields)

    snapshot_info = ensure_today_snapshot()

    file_id = FILE_IDS["qa_report"]
    baseline_modified = fetch_current_drive_modified(file_id)

    row_idx = get_bug_row_index_by_id(bug_id)
    if row_idx is None:
        raise SafetyError(f"BUG '{bug_id}' не знайдено.", kind="not_found")

    if not force_overwrite:
        check_drive_unchanged(file_id, baseline_modified)

    _raw, wb, ws, header_map = _open_workbook_writable()

    id_col = header_map.get("ID")
    if id_col is None:
        raise SafetyError("Колонка 'ID' не знайдена у header.", kind="header_missing")
    actual_id = ws.cell(row=row_idx, column=id_col).value
    if actual_id is None or str(actual_id).strip().upper() != bug_id.strip().upper():
        raise SafetyError(
            f"Рядок {row_idx} більше не містить {bug_id} (там '{actual_id}'). "
            f"Файл змінився між read і write — повтори операцію.",
            kind="row_mismatch",
        )

    updated_fields = []
    for field, value in fields.items():
        col_idx = header_map.get(field)
        if col_idx is None:
            raise SafetyError(
                f"Колонка '{field}' не знайдена у header.",
                kind="header_missing",
            )
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        if isinstance(value, str) and "\n" in value:
            from openpyxl.styles import Alignment
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical=existing.vertical or "top",
                wrap_text=True,
            )
        updated_fields.append(field)

    upload_meta = _save_and_upload(wb)
    clear_cache(file_id)

    log_write(
        tool="update_bug",
        record_id=bug_id,
        file_key="qa_report",
        fields_changed=updated_fields,
        force_overwrite=force_overwrite,
        row=row_idx,
        drive_modified_after=upload_meta.get("modifiedTime"),
    )

    return {
        "id": bug_id,
        "row": row_idx,
        "updated_fields": updated_fields,
        "snapshot": snapshot_info,
        "drive_modified_after": upload_meta.get("modifiedTime"),
        "force_overwrite_used": force_overwrite,
    }


def create_bug(fields: dict, use_id: str | None = None, force_overwrite: bool = False) -> dict:
    """
    Створює новий BUG.
    Якщо use_id переданий — використовує цей ID (після валідації і перевірки вільності),
    інакше — присвоює через next_bug_id.
    Запис завжди йде у КІНЕЦЬ файлу.
    """
    _validate_create_fields(fields)

    # Швидка валідація формату use_id ДО будь-яких Drive-операцій
    if use_id is not None:
        use_id = use_id.strip().upper()
        if not is_valid_bug_id_format(use_id):
            raise SafetyError(
                f"Невалідний формат use_id='{use_id}'. Очікую BUG-NNN або BUG-NNNN.",
                kind="invalid_id_format",
            )

    snapshot_info = ensure_today_snapshot()

    file_id = FILE_IDS["qa_report"]
    baseline_modified = fetch_current_drive_modified(file_id)

    if use_id is None:
        new_id = get_next_bug_id()
    else:
        new_id = use_id

    if not force_overwrite:
        check_drive_unchanged(file_id, baseline_modified)

    _raw, wb, ws, header_map = _open_workbook_writable()

    id_col = header_map.get("ID")
    if id_col is None:
        raise SafetyError("Колонка 'ID' не знайдена.", kind="header_missing")

    # Перевірка вільності ID на вже завантаженій книзі.
    new_id_upper = new_id.strip().upper()
    opis_col = header_map.get("Опис")
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
        cell = row[id_col - 1]
        if cell.value is None:
            continue
        if str(cell.value).strip().upper() == new_id_upper:
            occupant_preview = None
            if opis_col is not None:
                opis_val = ws.cell(row=cell.row, column=opis_col).value
                if opis_val is not None:
                    opis_str = str(opis_val)
                    occupant_preview = opis_str[:100] + ("..." if len(opis_str) > 100 else "")
            if use_id is not None:
                raise SafetyError(
                    f"ID {new_id} вже зайнятий (Опис: '{occupant_preview}'). "
                    f"Або вибери інший вільний ID, або викликай create без use_id.",
                    kind="id_collision",
                )
            raise SafetyError(
                f"ID {new_id} зайнятий — race condition (Опис: '{occupant_preview}').",
                kind="id_collision",
            )

    last_row = HEADER_ROW
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
        if all(c.value is None for c in row):
            continue
        last_row = row[0].row
    new_row = last_row + 1 if last_row >= DATA_START_ROW else DATA_START_ROW

    if last_row >= DATA_START_ROW:
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=last_row, column=col_idx)
            dst = ws.cell(row=new_row, column=col_idx)
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

    full_fields = {"ID": new_id, **fields}
    written_fields = []
    for field, value in full_fields.items():
        col_idx = header_map.get(field)
        if col_idx is None:
            raise SafetyError(
                f"Колонка '{field}' не знайдена у header.",
                kind="header_missing",
            )
        cell = ws.cell(row=new_row, column=col_idx)
        cell.value = value
        if isinstance(value, str) and "\n" in value:
            from openpyxl.styles import Alignment
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical=existing.vertical or "top",
                wrap_text=True,
            )
        written_fields.append(field)

    upload_meta = _save_and_upload(wb)
    clear_cache(file_id)

    log_write(
        tool="create_bug",
        record_id=new_id,
        file_key="qa_report",
        fields_changed=written_fields,
        force_overwrite=force_overwrite,
        row=new_row,
        drive_modified_after=upload_meta.get("modifiedTime"),
        fields_preview=full_fields,
    )

    return {
        "id": new_id,
        "row": new_row,
        "written_fields": written_fields,
        "snapshot": snapshot_info,
        "drive_modified_after": upload_meta.get("modifiedTime"),
        "force_overwrite_used": force_overwrite,
    }
