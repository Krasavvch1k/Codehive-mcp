"""
Worqen Workspace — write tools для gsheet і xlsx.

Шар обгорток над shared.sheets_client (gsheet через Sheets API) і
shared.drive_client.upload_file_content + openpyxl (xlsx).

1. Resolve query → file через ws_reader.resolve (gsheet/xlsx тільки)
2. Blacklist check (WRITE_BLACKLIST_FILE_IDS / NAME_SUBSTRINGS / FOLDERS)
3. Snapshot перед write (через safety.ensure_today_snapshot — auto, раз/день)
4. Drive-unchanged check (force_overwrite=True щоб обійти)
5. Маршрутизація: gsheet → sheets_client, xlsx → openpyxl + reupload
6. dry_run: повертає preview без write
7. Logging через writes_log.log_doc_write

Операції:
- update_cell(query, sheet, cell, value)   — одна комірка
- update_range(query, sheet, range, values) — 2D-блок
- append_row(query, sheet, values)         — додавання в кінець таблиці
"""

import io
import logging
import re
from copy import copy
from typing import Any, Optional

from googleapiclient.errors import HttpError
from openpyxl import load_workbook

from shared.drive_client import (
    download_file,
    fetch_current_drive_modified,
    upload_file_content,
)
from shared.safety import SafetyError, check_drive_unchanged
from shared.sheets_client import (
    append_values as _gsheet_append_values,
    get_spreadsheet_meta as _gsheet_meta,
    update_values as _gsheet_update_values,
)
from shared.writes_log import log_doc_write

from projects.worqen.config import (
    WRITE_BLACKLIST_FILE_IDS,
    WRITE_BLACKLIST_FOLDERS,
    WRITE_BLACKLIST_NAME_SUBSTRINGS,
    XLSX_MIME,
)
from projects.worqen.safety import ensure_today_snapshot
from projects.worqen.ws_reader import resolve

logger = logging.getLogger(__name__)


_WRITABLE_SHEET_KINDS = ("gsheet", "xlsx")
PROJECT = "worqen"

# A1 шаблон: одна комірка ("A1", "BC42") або діапазон ("A1:B10")
_A1_CELL_RE = re.compile(r"^[A-Z]+\d+$")
_A1_RANGE_RE = re.compile(r"^[A-Z]+\d+:[A-Z]+\d+$")


# ----- blacklist (дзеркало з ws_writer; винесу у shared safety пізніше якщо буде дубль) -----

def _check_blacklist(meta: dict) -> None:
    file_id = meta["id"]
    name = meta.get("name", "")
    parent_id = meta.get("parent_folder_id", "")

    if file_id in WRITE_BLACKLIST_FILE_IDS:
        raise SafetyError(
            f"Файл '{name}' (id={file_id}) у WRITE_BLACKLIST_FILE_IDS. "
            f"Для US/BUG використовуй worqen_update_user_story / worqen_update_bug.",
            kind="blacklisted_file_id",
        )

    if parent_id and parent_id in WRITE_BLACKLIST_FOLDERS:
        raise SafetyError(
            f"Папка батьківська (id={parent_id}) у WRITE_BLACKLIST_FOLDERS.",
            kind="blacklisted_folder",
        )

    name_lower = name.lower()
    for substr in WRITE_BLACKLIST_NAME_SUBSTRINGS:
        if substr.lower() in name_lower:
            raise SafetyError(
                f"Назва '{name}' містить '{substr}' (WRITE_BLACKLIST_NAME_SUBSTRINGS).",
                kind="blacklisted_name",
            )


def _resolve_writable_sheet(query: str) -> dict:
    """Резолвить query до gsheet/xlsx + blacklist check."""
    meta = resolve(query, kind_filter=_WRITABLE_SHEET_KINDS)
    _check_blacklist(meta)
    return meta


# ----- A1 helpers -----

def _validate_a1_cell(cell: str) -> None:
    if not _A1_CELL_RE.match(cell.upper()):
        raise ValueError(
            f"Невалідна A1-комірка: '{cell}'. Очікую формат 'A1', 'BC42'."
        )


def _validate_a1_range(range_str: str) -> None:
    if not _A1_RANGE_RE.match(range_str.upper()):
        raise ValueError(
            f"Невалідний A1-діапазон: '{range_str}'. "
            f"Очікую формат 'A1:B10'."
        )


def _build_full_range(sheet: str, a1: str) -> str:
    """
    Складає повний range для Sheets API: 'SheetName!A1' або "'My Sheet'!A1:B10".

    Назву у лапки беремо якщо є пробіл, апостроф або початок з цифри.
    """
    needs_quote = (
        " " in sheet
        or "'" in sheet
        or (sheet and sheet[0].isdigit())
    )
    if needs_quote:
        # Екрануємо одинарні лапки всередині (Sheets вимагає подвоєння)
        escaped = sheet.replace("'", "''")
        return f"'{escaped}'!{a1}"
    return f"{sheet}!{a1}"


def _a1_cell_to_col_row(cell: str) -> tuple[int, int]:
    """A1 → (col_index_1based, row_index_1based). 'A1' → (1, 1), 'AB10' → (28, 10)."""
    cell = cell.upper()
    m = re.match(r"^([A-Z]+)(\d+)$", cell)
    if not m:
        raise ValueError(f"Невалідна A1-комірка: '{cell}'")
    letters, digits = m.group(1), m.group(2)
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return col, int(digits)


# ----- xlsx helpers -----

def _open_xlsx_workbook(file_id: str) -> tuple[Any, Any]:
    """
    Завантажує xlsx з Drive (force_refresh) і відкриває у write-mode.

    Повертає (workbook, raw_bytes). Caller серіалізує і upload-ить.
    """
    raw = download_file(file_id, fmt="xlsx", force_refresh=True)
    wb = load_workbook(io.BytesIO(raw))  # default read_only=False
    return wb, raw


def _save_xlsx_and_upload(wb: Any, file_id: str) -> dict:
    """Серіалізує і заливає назад."""
    out = io.BytesIO()
    wb.save(out)
    return upload_file_content(file_id, out.getvalue(), XLSX_MIME)


def _resolve_xlsx_sheet(wb: Any, sheet_name: str) -> Any:
    """Повертає worksheet або кидає SafetyError якщо нема."""
    if sheet_name not in wb.sheetnames:
        raise SafetyError(
            f"Sheet '{sheet_name}' не знайдено. Доступні: {wb.sheetnames}",
            kind="sheet_missing",
        )
    return wb[sheet_name]


def _xlsx_find_last_data_row(ws: Any) -> int:
    """Знаходить номер останнього non-empty рядка. Якщо порожній — 0."""
    last = 0
    for row in ws.iter_rows(values_only=False):
        if all(c.value is None for c in row):
            continue
        last = row[0].row
    return last


def _xlsx_copy_row_style(ws: Any, src_row: int, dst_row: int) -> None:
    """Копіює стиль з src_row у dst_row по всіх колонках."""
    if src_row < 1:
        return
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col_idx)
        dst = ws.cell(row=dst_row, column=col_idx)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


# ----- public API: update_cell -----

def ws_update_cell(
    query: str,
    sheet: str,
    cell: str,
    value: Any,
    dry_run: bool = False,
    force_overwrite: bool = False,
) -> dict[str, Any]:
    """
    Перезаписує ОДНУ комірку.

    Args:
        query: substring у name або повний Drive ID файла.
        sheet: назва аркуша (рядково, як у файлі).
        cell: A1-нотація однієї комірки ('A1', 'BC42').
        value: нове значення (str/int/float/bool/None).
        dry_run: preview без запису.
        force_overwrite: пропустити drive-unchanged check.

    Returns:
        dict з ok / error / preview.
    """
    try:
        _validate_a1_cell(cell)
    except ValueError as e:
        return {"error": str(e), "kind": "invalid_a1"}

    try:
        meta = _resolve_writable_sheet(query)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    if dry_run:
        return {
            "operation": "update_cell",
            "file_id": meta["id"],
            "file_name": meta["name"],
            "kind": meta["kind"],
            "path": meta.get("path"),
            "sheet": sheet,
            "cell": cell.upper(),
            "value": value,
            "would_write": True,
            "dry_run": True,
        }

    snapshot_info = ensure_today_snapshot()
    file_id = meta["id"]
    baseline_modified = fetch_current_drive_modified(file_id)

    # drive sync check
    if not force_overwrite:
        try:
            check_drive_unchanged(file_id, baseline_modified)
        except SafetyError as e:
            return {
                "error": str(e),
                "kind": e.kind,
                "hint": "Передай force_overwrite=True щоб ігнорувати.",
            }

    try:
        if meta["kind"] == "gsheet":
            full_range = _build_full_range(sheet, cell.upper())
            api_resp = _gsheet_update_values(
                spreadsheet_id=file_id,
                range_str=full_range,
                values=[[value]],
            )
            modified_after = fetch_current_drive_modified(file_id)
            result_payload = {
                "updated_range": api_resp.get("updatedRange"),
                "updated_cells": api_resp.get("updatedCells"),
            }
        else:  # xlsx
            wb, _ = _open_xlsx_workbook(file_id)
            ws = _resolve_xlsx_sheet(wb, sheet)
            col, row = _a1_cell_to_col_row(cell)
            ws.cell(row=row, column=col, value=value)
            upload_meta = _save_xlsx_and_upload(wb, file_id)
            modified_after = upload_meta.get("modifiedTime")
            result_payload = {"row": row, "col": col}

    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except HttpError as e:
        return {"error": f"API failed: {e}", "kind": "api_error"}
    except Exception as e:
        return {"error": f"unexpected: {e}", "kind": "unexpected"}

    try:
        log_doc_write(
            project=PROJECT,
            tool="ws_update_cell",
            file_id=file_id,
            file_name=meta["name"],
            payload={
                "kind": meta["kind"],
                "sheet": sheet,
                "cell": cell.upper(),
                "value": str(value) if value is not None else None,
                "modified_before": baseline_modified,
                "modified_after": modified_after,
                "force_overwrite": force_overwrite,
            },
        )
    except Exception as e:
        logger.warning("writes_log failed for ws_update_cell: %s", e)

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": meta["name"],
        "kind": meta["kind"],
        "sheet": sheet,
        "cell": cell.upper(),
        "value": value,
        "snapshot": snapshot_info,
        "modified_before": baseline_modified,
        "modified_after": modified_after,
        "force_overwrite_used": force_overwrite,
        **result_payload,
    }


# ----- public API: update_range -----

def ws_update_range(
    query: str,
    sheet: str,
    range_a1: str,
    values: list[list],
    dry_run: bool = False,
    force_overwrite: bool = False,
) -> dict[str, Any]:
    """
    Перезаписує прямокутний діапазон значень.

    Args:
        query: substring у name або Drive ID.
        sheet: назва аркуша.
        range_a1: A1-діапазон ('A1:B10').
        values: 2D-список рядок-стовпець. Має співпадати розміром з range_a1
            (або менше — тоді хвіст не оновиться, але краще явно).
        dry_run: preview без запису.
        force_overwrite: пропустити drive-unchanged check.

    Returns:
        dict з ok / error / preview.
    """
    try:
        _validate_a1_range(range_a1)
    except ValueError as e:
        return {"error": str(e), "kind": "invalid_a1"}

    if not values or not all(isinstance(r, list) for r in values):
        return {
            "error": "values має бути 2D-list (list of lists)",
            "kind": "invalid_values",
        }

    try:
        meta = _resolve_writable_sheet(query)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    rows_count = len(values)
    cols_count = max(len(r) for r in values) if values else 0

    if dry_run:
        return {
            "operation": "update_range",
            "file_id": meta["id"],
            "file_name": meta["name"],
            "kind": meta["kind"],
            "path": meta.get("path"),
            "sheet": sheet,
            "range": range_a1.upper(),
            "rows": rows_count,
            "cols": cols_count,
            "values_preview": [r[:5] for r in values[:5]],  # 5x5 first
            "would_write": True,
            "dry_run": True,
        }

    snapshot_info = ensure_today_snapshot()
    file_id = meta["id"]
    baseline_modified = fetch_current_drive_modified(file_id)

    if not force_overwrite:
        try:
            check_drive_unchanged(file_id, baseline_modified)
        except SafetyError as e:
            return {
                "error": str(e),
                "kind": e.kind,
                "hint": "Передай force_overwrite=True щоб ігнорувати.",
            }

    try:
        if meta["kind"] == "gsheet":
            full_range = _build_full_range(sheet, range_a1.upper())
            api_resp = _gsheet_update_values(
                spreadsheet_id=file_id,
                range_str=full_range,
                values=values,
            )
            modified_after = fetch_current_drive_modified(file_id)
            result_payload = {
                "updated_range": api_resp.get("updatedRange"),
                "updated_cells": api_resp.get("updatedCells"),
            }
        else:  # xlsx
            wb, _ = _open_xlsx_workbook(file_id)
            ws = _resolve_xlsx_sheet(wb, sheet)
            # Парсимо start cell з range_a1
            start_cell, _end_cell = range_a1.upper().split(":")
            start_col, start_row = _a1_cell_to_col_row(start_cell)
            for r_offset, row_values in enumerate(values):
                for c_offset, val in enumerate(row_values):
                    ws.cell(
                        row=start_row + r_offset,
                        column=start_col + c_offset,
                        value=val,
                    )
            upload_meta = _save_xlsx_and_upload(wb, file_id)
            modified_after = upload_meta.get("modifiedTime")
            result_payload = {
                "start_row": start_row,
                "start_col": start_col,
                "rows_written": rows_count,
                "cols_written": cols_count,
            }

    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except HttpError as e:
        return {"error": f"API failed: {e}", "kind": "api_error"}
    except Exception as e:
        return {"error": f"unexpected: {e}", "kind": "unexpected"}

    try:
        log_doc_write(
            project=PROJECT,
            tool="ws_update_range",
            file_id=file_id,
            file_name=meta["name"],
            payload={
                "kind": meta["kind"],
                "sheet": sheet,
                "range": range_a1.upper(),
                "rows": rows_count,
                "cols": cols_count,
                "modified_before": baseline_modified,
                "modified_after": modified_after,
                "force_overwrite": force_overwrite,
            },
        )
    except Exception as e:
        logger.warning("writes_log failed for ws_update_range: %s", e)

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": meta["name"],
        "kind": meta["kind"],
        "sheet": sheet,
        "range": range_a1.upper(),
        "rows": rows_count,
        "cols": cols_count,
        "snapshot": snapshot_info,
        "modified_before": baseline_modified,
        "modified_after": modified_after,
        "force_overwrite_used": force_overwrite,
        **result_payload,
    }


# ----- public API: append_row -----

def ws_append_row(
    query: str,
    sheet: str,
    values: list,
    copy_style_from_last: bool = True,
    dry_run: bool = False,
    force_overwrite: bool = False,
) -> dict[str, Any]:
    """
    Додає один рядок у кінець таблиці.

    Для gsheet — через values.append (Sheets API сам шукає кінець таблиці).
    Для xlsx — знаходимо last_data_row + 1, пишемо туди, опційно копіюємо стиль.

    Args:
        query: substring у name або Drive ID.
        sheet: назва аркуша.
        values: 1D-list значень (не 2D). Один рядок.
        copy_style_from_last: для xlsx — копіювати стиль з попереднього рядка
            (font/fill/border/alignment).
        dry_run: preview без запису.
        force_overwrite: bypass drive-unchanged check.

    Returns:
        dict з ok / error / preview.
    """
    if not isinstance(values, list) or any(isinstance(v, list) for v in values):
        return {
            "error": "values для append_row має бути 1D-list, не 2D.",
            "kind": "invalid_values",
        }

    try:
        meta = _resolve_writable_sheet(query)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    if dry_run:
        return {
            "operation": "append_row",
            "file_id": meta["id"],
            "file_name": meta["name"],
            "kind": meta["kind"],
            "path": meta.get("path"),
            "sheet": sheet,
            "row_values": values,
            "row_length": len(values),
            "copy_style_from_last": copy_style_from_last,
            "would_write": True,
            "dry_run": True,
        }

    snapshot_info = ensure_today_snapshot()
    file_id = meta["id"]
    baseline_modified = fetch_current_drive_modified(file_id)

    if not force_overwrite:
        try:
            check_drive_unchanged(file_id, baseline_modified)
        except SafetyError as e:
            return {
                "error": str(e),
                "kind": e.kind,
                "hint": "Передай force_overwrite=True щоб ігнорувати.",
            }

    appended_row_idx: Optional[int] = None

    try:
        if meta["kind"] == "gsheet":
            # Для append range = sheet name (Sheets знайде кінець сам)
            full_range = _build_full_range(sheet, "A:A")
            api_resp = _gsheet_append_values(
                spreadsheet_id=file_id,
                range_str=full_range,
                values=[values],
            )
            modified_after = fetch_current_drive_modified(file_id)
            updates = api_resp.get("updates", {})
            result_payload = {
                "updated_range": updates.get("updatedRange"),
                "updated_cells": updates.get("updatedCells"),
            }
        else:  # xlsx
            wb, _ = _open_xlsx_workbook(file_id)
            ws = _resolve_xlsx_sheet(wb, sheet)
            last_row = _xlsx_find_last_data_row(ws)
            new_row = last_row + 1 if last_row >= 1 else 1

            if copy_style_from_last and last_row >= 1:
                _xlsx_copy_row_style(ws, last_row, new_row)

            for c_offset, val in enumerate(values, start=1):
                ws.cell(row=new_row, column=c_offset, value=val)

            upload_meta = _save_xlsx_and_upload(wb, file_id)
            modified_after = upload_meta.get("modifiedTime")
            appended_row_idx = new_row
            result_payload = {
                "row": new_row,
                "cols_written": len(values),
                "style_copied": copy_style_from_last and last_row >= 1,
            }

    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except HttpError as e:
        return {"error": f"API failed: {e}", "kind": "api_error"}
    except Exception as e:
        return {"error": f"unexpected: {e}", "kind": "unexpected"}

    try:
        log_doc_write(
            project=PROJECT,
            tool="ws_append_row",
            file_id=file_id,
            file_name=meta["name"],
            payload={
                "kind": meta["kind"],
                "sheet": sheet,
                "row_length": len(values),
                "appended_row_idx": appended_row_idx,
                "modified_before": baseline_modified,
                "modified_after": modified_after,
                "force_overwrite": force_overwrite,
            },
        )
    except Exception as e:
        logger.warning("writes_log failed for ws_append_row: %s", e)

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": meta["name"],
        "kind": meta["kind"],
        "sheet": sheet,
        "row_length": len(values),
        "snapshot": snapshot_info,
        "modified_before": baseline_modified,
        "modified_after": modified_after,
        "force_overwrite_used": force_overwrite,
        **result_payload,
    }
