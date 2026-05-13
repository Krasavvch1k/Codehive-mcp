"""
Worqen Workspace — читання spreadsheet-файлів (gsheet і xlsx).

gsheet → Google Sheets API (shared.sheets_client)
xlsx   → Drive API download + openpyxl

Обидва шляхи повертають однорідну структуру: список аркушів з даними і
метаданими (rows / cols / range_read).
"""

import io
from typing import Optional

from openpyxl import load_workbook

from projects.worqen.ws_reader import list_all, resolve
from shared.drive_client import download_file
from shared.sheets_client import get_sheet_values, get_spreadsheet_meta


# Скільки рядків/колонок повертати за замовчуванням — щоб не вантажити
# мегабайти даних з великих таблиць. Користувач може запросити більше через
# параметри limit_rows / limit_cols.
DEFAULT_LIMIT_ROWS = 200
DEFAULT_LIMIT_COLS = 50
MAX_LIMIT_ROWS = 5000
MAX_LIMIT_COLS = 200


def _normalize_row(row: list, cols: int) -> list:
    """Доповнює рядок до cols пустими рядками (Sheets API обрізає trailing empty)."""
    if len(row) >= cols:
        return [_to_str(c) for c in row[:cols]]
    return [_to_str(c) for c in row] + [""] * (cols - len(row))


def _to_str(value) -> str:
    """Зводить будь-яке значення до str для однорідного виводу."""
    if value is None:
        return ""
    return str(value)


def _build_markdown_table(rows: list[list[str]], has_header: bool = True) -> str:
    """Форматує 2D-список у markdown-таблицю."""
    if not rows:
        return ""
    n_cols = max(len(r) for r in rows)
    normalized = [_normalize_row(r, n_cols) for r in rows]

    lines: list[str] = []
    if has_header and normalized:
        header = normalized[0]
        lines.append(" | ".join(header))
        lines.append(" | ".join(["---"] * n_cols))
        body = normalized[1:]
    else:
        body = normalized
    for r in body:
        # Замінюємо переноси у значеннях на пробіли (markdown table inline)
        r_clean = [v.replace("\n", " ").replace("|", "\\|") for v in r]
        lines.append(" | ".join(r_clean))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# gsheet через Sheets API
# ---------------------------------------------------------------------------


def _read_gsheet(
    file_id: str,
    sheet_name: Optional[str],
    limit_rows: int,
    limit_cols: int,
    force_refresh: bool,
    as_markdown: bool,
) -> dict:
    meta = get_spreadsheet_meta(file_id, force_refresh=force_refresh)

    sheets = meta["sheets"]
    if not sheets:
        return {
            "id": file_id,
            "title": meta.get("title"),
            "sheets_meta": [],
            "sheets_data": [],
            "warning": "spreadsheet не містить аркушів",
        }

    # Які аркуші читаємо
    if sheet_name:
        target = [s for s in sheets if s["title"] == sheet_name]
        if not target:
            available = ", ".join(s["title"] for s in sheets)
            raise ValueError(
                f"Аркуш '{sheet_name}' не знайдено. Доступні: {available}"
            )
    else:
        target = sheets  # читаємо всі

    sheets_data: list[dict] = []
    for s in target:
        title = s["title"]
        row_cap = min(s["row_count"], limit_rows)
        col_cap = min(s["col_count"], limit_cols)
        # Будуємо range: 'Title'!A1:<col_letter><row>
        # Sheets API розуміє і просту назву — тоді поверне весь аркуш,
        # але з cap по рядках/колонках безпечніше:
        col_letter = _col_index_to_letter(col_cap)
        range_str = f"'{title}'!A1:{col_letter}{row_cap}"
        try:
            values = get_sheet_values(
                file_id,
                range_str,
                force_refresh=force_refresh,
            )
        except Exception as e:
            sheets_data.append({
                "sheet_id": s["sheet_id"],
                "title": title,
                "error": str(e),
            })
            continue

        normalized = [[_to_str(c) for c in row] for row in values]
        # Sheets API зазвичай обрізає trailing empty, але буває з заповненими
        # лише першими N рядками теж повертає весь діапазон. Для консистентності
        # з xlsx робимо те саме обрізання.
        while normalized and all(c.strip() == "" for c in normalized[-1]):
            normalized.pop()

        # Реальна ширина даних
        max_used_cols = 0
        for r in normalized:
            last_nonempty = 0
            for i, c in enumerate(r):
                if c.strip():
                    last_nonempty = i + 1
            if last_nonempty > max_used_cols:
                max_used_cols = last_nonempty
        if max_used_cols > 0:
            normalized = [r[:max_used_cols] for r in normalized]

        entry = {
            "sheet_id": s["sheet_id"],
            "title": title,
            "range_read": range_str,
            "rows_returned": len(normalized),
            "cols_returned": max_used_cols,
            "values": normalized,
        }
        if as_markdown:
            entry["markdown"] = _build_markdown_table(normalized, has_header=True)
        sheets_data.append(entry)

    return {
        "id": file_id,
        "kind": "gsheet",
        "title": meta.get("title"),
        "sheets_meta": sheets,
        "sheets_data": sheets_data,
        "force_refresh": force_refresh,
    }


def _col_index_to_letter(n: int) -> str:
    """1 -> A, 26 -> Z, 27 -> AA. n клонується у мінімум 1."""
    n = max(1, n)
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ---------------------------------------------------------------------------
# xlsx через openpyxl
# ---------------------------------------------------------------------------


def _read_xlsx(
    file_id: str,
    sheet_name: Optional[str],
    limit_rows: int,
    limit_cols: int,
    force_refresh: bool,
    as_markdown: bool,
) -> dict:
    content = download_file(file_id, fmt="xlsx", force_refresh=force_refresh)
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # У read_only режимі openpyxl інколи повертає max_row/max_column=None (0)
    # до першої ітерації. Не покладаємось на них — використовуємо limit як cap
    # і покладаємось на trim-trailing-empty.

    all_sheets_meta = [
        {
            "title": name,
            "index": i,
        }
        for i, name in enumerate(wb.sheetnames)
    ]

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"Аркуш '{sheet_name}' не знайдено. Доступні: {available}"
            )
        target_names = [sheet_name]
    else:
        target_names = wb.sheetnames

    sheets_data: list[dict] = []
    for name in target_names:
        ws = wb[name]
        # max_row/max_col у read_only — ненадійні (часто 0/None). Беремо як hint
        # тільки якщо вони дають менше за limit_rows; інакше використовуємо limit
        # як cap і обрізаємо trailing-empty у кінці.
        hinted_rows = ws.max_row or 0
        hinted_cols = ws.max_column or 0
        row_cap = min(hinted_rows, limit_rows) if hinted_rows > 0 else limit_rows
        col_cap = min(hinted_cols, limit_cols) if hinted_cols > 0 else limit_cols

        rows: list[list[str]] = []
        for row in ws.iter_rows(
            min_row=1, max_row=row_cap, max_col=col_cap, values_only=True
        ):
            rows.append([_to_str(c) for c in row])

        # Обрізаємо trailing empty rows (усі комірки порожні)
        while rows and all(c.strip() == "" for c in rows[-1]):
            rows.pop()

        # Реальна ширина даних (по кожному рядку) — стискаємо до max_used_cols
        max_used_cols = 0
        for r in rows:
            last_nonempty = 0
            for i, c in enumerate(r):
                if c.strip():
                    last_nonempty = i + 1
            if last_nonempty > max_used_cols:
                max_used_cols = last_nonempty
        if max_used_cols > 0:
            rows = [r[:max_used_cols] for r in rows]

        entry = {
            "title": name,
            "range_read": f"A1:{_col_index_to_letter(col_cap)}{row_cap}",
            "rows_returned": len(rows),
            "cols_returned": max_used_cols,
            "values": rows,
        }
        if as_markdown:
            entry["markdown"] = _build_markdown_table(rows, has_header=True)
        sheets_data.append(entry)

    wb.close()

    return {
        "id": file_id,
        "kind": "xlsx",
        "title": None,  # xlsx нема title окремо від filename
        "sheets_meta": all_sheets_meta,
        "sheets_data": sheets_data,
        "force_refresh": force_refresh,
    }


# ---------------------------------------------------------------------------
# Public: read_sheet
# ---------------------------------------------------------------------------


def read_sheet(
    query: str,
    sheet_name: Optional[str] = None,
    limit_rows: int = DEFAULT_LIMIT_ROWS,
    limit_cols: int = DEFAULT_LIMIT_COLS,
    force_refresh: bool = False,
    as_markdown: bool = True,
) -> dict:
    """
    Читає gsheet або xlsx файл у workspace.

    query: name substring або повний Drive ID.
    sheet_name: якщо вказано — читає тільки цей аркуш. Інакше — всі.
    limit_rows / limit_cols: обмеження щоб не тягнути мегабайти.
    force_refresh: обходить TTL-кеш.
    as_markdown: додавати markdown-репрезентацію кожного аркуша.
    """
    limit_rows = max(1, min(limit_rows, MAX_LIMIT_ROWS))
    limit_cols = max(1, min(limit_cols, MAX_LIMIT_COLS))

    all_data = list_all()
    candidates = [d for d in all_data["items"] if d["kind"] in ("gsheet", "xlsx")]
    meta = resolve(query, candidates)

    common_meta = {
        "name": meta["name"],
        "path": meta.get("path"),
        "modified": meta.get("modified"),
        "parent_folder_id": meta.get("parent_folder_id"),
        "parent_folder_name": meta.get("parent_folder_name"),
    }

    if meta["kind"] == "gsheet":
        result = _read_gsheet(
            meta["id"], sheet_name, limit_rows, limit_cols, force_refresh, as_markdown,
        )
    else:
        result = _read_xlsx(
            meta["id"], sheet_name, limit_rows, limit_cols, force_refresh, as_markdown,
        )

    result.update(common_meta)
    return result
