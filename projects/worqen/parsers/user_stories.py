"""
Парсер Worqen_User_Stories_.xlsx з tier-системою.
"""

import io
import re
from typing import Optional
from openpyxl import load_workbook

from shared.drive_client import download_file
from projects.worqen.config import FILE_IDS, SCAN_TITLE_TRUNCATE


SHEET_NAME = "User Stories"

ALL_COLUMNS = [
    "Epic", "ID", "Title", "User Story", "Status", "Priority", "Version",
    "Est. day", "Acceptance Criteria", "Edge Cases", "Dependencies",
    "Notes", "Related Decisions",
]

SCAN_FIELDS = ["Epic", "ID", "Title", "Status", "Priority", "Version"]
AUDIT_FIELDS = SCAN_FIELDS + [
    "User Story", "Acceptance Criteria", "Edge Cases",
    "Dependencies", "Notes", "Related Decisions",
]


def _load_all_stories() -> list[dict]:
    content = download_file(FILE_IDS["user_stories"])
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found")
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    stories = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        story = {}
        for i, header in enumerate(headers):
            if header is None:
                continue
            value = row[i] if i < len(row) else None
            story[header] = None if value is None else str(value)
        if not story.get("ID"):
            continue
        stories.append(story)
    return stories


def _truncate(value, limit=SCAN_TITLE_TRUNCATE):
    if value is None:
        return None
    if len(value) <= limit:
        return value
    return value[: limit - 3] + "..."


def _apply_tier(story, tier):
    if tier == "full":
        return story
    if tier == "audit":
        return {k: story.get(k) for k in AUDIT_FIELDS if k in story}
    result = {k: story.get(k) for k in SCAN_FIELDS if k in story}
    if "Title" in result and result["Title"]:
        result["Title"] = _truncate(result["Title"])
    return result


def list_stories(epic=None, status=None, priority=None, version=None, search=None, tier="scan"):
    stories = _load_all_stories()
    result = []
    for s in stories:
        if epic and epic.upper() not in (s.get("Epic") or "").upper():
            continue
        if status and (s.get("Status") or "").lower() != status.lower():
            continue
        if priority and (s.get("Priority") or "").lower() != priority.lower():
            continue
        if version and (s.get("Version") or "").lower() != version.lower():
            continue
        if search:
            haystack = " ".join(
                str(s.get(k) or "")
                for k in ("Title", "User Story", "Notes", "Acceptance Criteria",
                         "Edge Cases", "Dependencies", "Related Decisions")
            ).lower()
            if search.lower() not in haystack:
                continue
        result.append(_apply_tier(s, tier))
    return result


def get_story(story_id):
    sid = story_id.strip().upper()
    for s in _load_all_stories():
        if (s.get("ID") or "").strip().upper() == sid:
            return s
    return None


def get_next_us_id():
    max_num = 0
    pattern = re.compile(r"^US-(\d+)$", re.IGNORECASE)
    for s in _load_all_stories():
        m = pattern.match((s.get("ID") or "").strip())
        if m:
            num = int(m.group(1))
            if num > max_num:
                max_num = num
    return f"US-{max_num + 1:03d}"


def list_epics_summary(level="minimal"):
    stories = _load_all_stories()
    summary = {}
    for s in stories:
        epic = s.get("Epic") or "(no epic)"
        if epic not in summary:
            summary[epic] = {
                "total": 0,
                "by_status": {},
                "by_priority": {},
                "by_version": {},
            }
        summary[epic]["total"] += 1
        st = s.get("Status") or "(no status)"
        summary[epic]["by_status"][st] = summary[epic]["by_status"].get(st, 0) + 1
        pr = s.get("Priority") or "(no priority)"
        summary[epic]["by_priority"][pr] = summary[epic]["by_priority"].get(pr, 0) + 1
        ver = s.get("Version") or "(no version)"
        summary[epic]["by_version"][ver] = summary[epic]["by_version"].get(ver, 0) + 1

    if level == "minimal":
        return {epic: data["total"] for epic, data in summary.items()}
    if level == "status":
        return {
            epic: {"total": data["total"], "by_status": data["by_status"]}
            for epic, data in summary.items()
        }
    return summary


def find_dependents(story_id, tier="scan"):
    sid = story_id.strip().upper()
    result = []
    for s in _load_all_stories():
        deps = (s.get("Dependencies") or "").upper()
        if sid in deps:
            result.append(_apply_tier(s, tier))
    return result


def get_raw_stories():
    return _load_all_stories()


def get_row_index_by_id(story_id: str) -> Optional[int]:
    """
    Повертає 1-based номер рядка в xlsx для вказаного US-ID.
    Header у row 1, дані з row 2.
    None якщо не знайдено.
    """
    sid = story_id.strip().upper()
    content = download_file(FILE_IDS["user_stories"])
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return None
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    headers = rows[0]
    try:
        id_col = list(headers).index("ID")
    except ValueError:
        return None
    for i, row in enumerate(rows[1:], start=2):
        if id_col >= len(row):
            continue
        cell = row[id_col]
        if cell is None:
            continue
        if str(cell).strip().upper() == sid:
            return i
    return None


def get_last_data_row_index() -> int:
    """
    Повертає 1-based номер ОСТАННЬОГО рядка з даними (для копіювання стилю
    при create нового рядка).
    Якщо даних немає — повертає 1 (header row).
    """
    content = download_file(FILE_IDS["user_stories"])
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    last = 1
    for i, row in enumerate(rows[1:], start=2):
        if all(c is None for c in row):
            continue
        last = i
    return last


def find_missing_us_ids() -> list[str]:
    """
    Повертає список пропусків у послідовності US-ID.
    Наприклад якщо є US-001, US-002, US-004 — повертає ['US-003'].
    Порядок зростаючий. Якщо пропусків немає — порожній список.
    """
    pattern = re.compile(r"^US-(\d+)$", re.IGNORECASE)
    existing_nums = set()
    max_num = 0
    for s in _load_all_stories():
        m = pattern.match((s.get("ID") or "").strip())
        if m:
            num = int(m.group(1))
            existing_nums.add(num)
            if num > max_num:
                max_num = num
    if max_num == 0:
        return []
    missing = []
    for n in range(1, max_num + 1):
        if n not in existing_nums:
            missing.append(f"US-{n:03d}")
    return missing


def is_us_id_free(story_id: str) -> tuple[bool, Optional[str]]:
    """
    Перевіряє чи вільний вказаний US-ID.
    Повертає (is_free, occupant_title).
    occupant_title — Title існуючого якщо ID зайнятий, інакше None.
    """
    sid = story_id.strip().upper()
    for s in _load_all_stories():
        if (s.get("ID") or "").strip().upper() == sid:
            return False, s.get("Title")
    return True, None


_US_ID_PATTERN = re.compile(r"^US-\d{3,4}$", re.IGNORECASE)


def is_valid_us_id_format(story_id: str) -> bool:
    """Перевірка формату: US-NNN або US-NNNN (3-4 цифри)."""
    if not story_id:
        return False
    return bool(_US_ID_PATTERN.match(story_id.strip()))
