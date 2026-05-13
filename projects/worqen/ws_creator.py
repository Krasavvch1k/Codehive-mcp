"""
Worqen Workspace — create tools для нових файлів/папок.

3 операції:
- ws_create_doc(name, parent_folder?, format=gdoc|docx, initial_content?)
- ws_create_sheet(name, parent_folder?, format=gsheet|xlsx, columns?)
- ws_create_folder(name, parent_folder?)

Поведінка:
- parent_folder необов'язковий — default корінь Worqen Drive (WORQEN_ROOT_FOLDER_ID)
- parent_folder приймає name (substring) або повний Drive ID
- Duplicate check: помилка якщо у parent_folder вже є файл/папка з такою назвою
  (можна обійти через allow_duplicate=True)
- Blacklist для create — тільки WRITE_BLACKLIST_FOLDERS (folder-level)
- dry_run: повертає preview без створення
- Logging через writes_log.log_doc_write

NB: gsheet через Sheets API spreadsheets.create кладе файл у root service
account-а — обхід через Drive API parents у файл-create. xlsx створюємо
через openpyxl + upload, gdoc — через Drive API create + опційне insert_text.
"""

import io
import logging
from typing import Any, Optional

from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import Workbook

from shared.doc_writers import create_doc_in_folder as _create_gdoc_in_folder
from shared.docx_writers import (
    DOCX_MIME,
    create_docx_in_folder as _create_docx_in_folder,
)
from shared.drive_client import _list_folder_children, get_service
from shared.safety import SafetyError
from shared.writes_log import log_doc_write

from projects.worqen.config import (
    GOOGLE_FOLDER_MIME,
    GOOGLE_SHEET_MIME,
    WORQEN_ROOT_FOLDER_ID,
    WRITE_BLACKLIST_FOLDERS,
    XLSX_MIME,
)
from projects.worqen.ws_reader import resolve

logger = logging.getLogger(__name__)


PROJECT = "worqen"

VALID_DOC_FORMATS = ("gdoc", "docx")
VALID_SHEET_FORMATS = ("gsheet", "xlsx")


# ----- parent folder resolution -----

def _resolve_parent_folder(parent_folder: Optional[str]) -> tuple[str, str]:
    """
    Резолвить parent_folder до (folder_id, folder_name).

    - None / "" → корінь Worqen Drive (WORQEN_ROOT_FOLDER_ID, name="Worqen")
    - повний Drive ID → точне співпадіння серед folder candidates
    - інакше substring у name (case-insensitive)

    Raises:
        ValueError якщо не знайдено / ambiguous.
        SafetyError якщо у WRITE_BLACKLIST_FOLDERS.
    """
    if not parent_folder:
        if not WORQEN_ROOT_FOLDER_ID:
            raise ValueError(
                "WORQEN_ROOT_FOLDER_ID не налаштовано у .env. "
                "Передай parent_folder явно."
            )
        if WORQEN_ROOT_FOLDER_ID in WRITE_BLACKLIST_FOLDERS:
            raise SafetyError(
                f"Корінь Worqen Drive ({WORQEN_ROOT_FOLDER_ID}) у "
                f"WRITE_BLACKLIST_FOLDERS — create заборонений.",
                kind="blacklisted_folder",
            )
        return WORQEN_ROOT_FOLDER_ID, "Worqen"

    folder = resolve(parent_folder, kind_filter=("folder",))

    if folder["id"] in WRITE_BLACKLIST_FOLDERS:
        raise SafetyError(
            f"Папка '{folder['name']}' (id={folder['id']}) у "
            f"WRITE_BLACKLIST_FOLDERS — create заборонений.",
            kind="blacklisted_folder",
        )

    return folder["id"], folder["name"]


def _check_duplicate(folder_id: str, name: str, allow_duplicate: bool) -> None:
    """
    Перевіряє чи у folder_id вже є файл/папка з такою назвою (case-insensitive).

    Raises:
        SafetyError якщо знайдено і allow_duplicate=False.
    """
    if allow_duplicate:
        return

    name_lower = name.lower()
    children = _list_folder_children(folder_id)
    for c in children:
        if c.get("name", "").lower() == name_lower:
            raise SafetyError(
                f"У папці вже є файл/папка '{c['name']}' (id={c['id']}). "
                f"Передай allow_duplicate=True щоб все одно створити.",
                kind="duplicate_name",
            )


# ----- xlsx creation helper -----

def _create_xlsx_in_folder(
    folder_id: str,
    folder_name: str,
    name: str,
    columns: Optional[list[str]],
) -> dict[str, Any]:
    """
    Створює новий xlsx у Drive folder з опційним header row.

    Returns dict з ok / error.
    """
    wb = Workbook()
    ws = wb.active
    if columns:
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

    out = io.BytesIO()
    wb.save(out)
    content_bytes = out.getvalue()

    upload_name = name if name.lower().endswith(".xlsx") else name + ".xlsx"

    service = get_service()
    media = MediaIoBaseUpload(
        io.BytesIO(content_bytes),
        mimetype=XLSX_MIME,
        resumable=False,
    )

    try:
        new_file = (
            service.files()
            .create(
                body={
                    "name": upload_name,
                    "mimeType": XLSX_MIME,
                    "parents": [folder_id],
                },
                media_body=media,
                fields="id, name, mimeType, parents, webViewLink, modifiedTime",
                supportsAllDrives=True,
            )
            .execute()
        )
    except HttpError as e:
        return {"error": f"Drive create failed: {e}", "kind": "api_error"}

    file_id = new_file["id"]
    web_view_link = new_file.get("webViewLink") or (
        f"https://drive.google.com/file/d/{file_id}/view"
    )

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": upload_name,
        "folder_id": folder_id,
        "folder_name": folder_name,
        "url": web_view_link,
        "modified_after": new_file.get("modifiedTime"),
        "columns_count": len(columns) if columns else 0,
    }


# ----- gsheet creation helper -----

def _create_gsheet_in_folder(
    folder_id: str,
    folder_name: str,
    name: str,
    columns: Optional[list[str]],
) -> dict[str, Any]:
    """
    Створює новий gsheet у Drive folder через Drive API + опційний header.

    Не використовуємо spreadsheets.create зі Sheets API, бо він кладе файл
    у root service account-а. Замість цього — Drive API files.create з
    mimeType=spreadsheet (Drive створить порожній gsheet) і потім (якщо
    columns передані) Sheets API values.update для header row.

    Returns dict з ok / error.
    """
    service = get_service()

    try:
        new_file = (
            service.files()
            .create(
                body={
                    "name": name,
                    "mimeType": GOOGLE_SHEET_MIME,
                    "parents": [folder_id],
                },
                fields="id, name, mimeType, parents, webViewLink, modifiedTime",
                supportsAllDrives=True,
            )
            .execute()
        )
    except HttpError as e:
        return {"error": f"Drive create failed: {e}", "kind": "api_error"}

    file_id = new_file["id"]
    web_view_link = new_file.get("webViewLink") or (
        f"https://docs.google.com/spreadsheets/d/{file_id}/edit"
    )

    columns_written = 0
    if columns:
        # Header у Sheet1!A1 — Sheets API називає default аркуш "Sheet1" /
        # локалізовано. Найбезпечніше — використати unqualified A1.
        from shared.sheets_client import update_values

        try:
            update_values(
                spreadsheet_id=file_id,
                range_str=f"A1:{_idx_to_letter(len(columns))}1",
                values=[columns],
            )
            columns_written = len(columns)
        except HttpError as e:
            return {
                "ok": True,
                "warning": (
                    f"Spreadsheet створений, але header row не записався: {e}. "
                    f"Спробуй ws_update_range з нужним діапазоном."
                ),
                "file_id": file_id,
                "file_name": name,
                "folder_id": folder_id,
                "folder_name": folder_name,
                "url": web_view_link,
                "modified_after": new_file.get("modifiedTime"),
                "columns_count": 0,
            }

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": name,
        "folder_id": folder_id,
        "folder_name": folder_name,
        "url": web_view_link,
        "modified_after": new_file.get("modifiedTime"),
        "columns_count": columns_written,
    }


def _idx_to_letter(n: int) -> str:
    """1-based column index → A1 letter (mirror of sheets_client._col_index_to_letter)."""
    if n < 1:
        raise ValueError(f"Column index must be >= 1, got {n}")
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# ----- public API: create_doc -----

def ws_create_doc(
    name: str,
    parent_folder: Optional[str] = None,
    format: str = "gdoc",
    initial_content: str = "",
    allow_duplicate: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Створює новий документ (gdoc або docx) у Worqen workspace.

    Args:
        name: назва нового файлу. Для docx суфікс .docx додається автоматично.
        parent_folder: name substring / Drive ID. Default — корінь Worqen Drive.
        format: 'gdoc' (default) або 'docx'.
        initial_content: опційний plain text (для docx — кожен '\\n' = новий
            параграф; для gdoc — звичайна вставка тексту).
        allow_duplicate: дозволити створення коли у папці вже є файл з такою
            назвою.
        dry_run: preview без створення.

    Returns:
        dict з ok / error / preview.
    """
    if format not in VALID_DOC_FORMATS:
        return {
            "error": f"format має бути одним з {VALID_DOC_FORMATS}, "
                     f"отримано: '{format}'",
            "kind": "invalid_format",
        }

    if not name or not name.strip():
        return {"error": "name не може бути порожнім", "kind": "invalid_name"}

    try:
        folder_id, folder_name = _resolve_parent_folder(parent_folder)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    if dry_run:
        return {
            "operation": "create_doc",
            "name": name,
            "format": format,
            "parent_folder_id": folder_id,
            "parent_folder_name": folder_name,
            "initial_content_length": len(initial_content),
            "allow_duplicate": allow_duplicate,
            "would_create": True,
            "dry_run": True,
            "note": (
                "Це preview без створення. Передай dry_run=False (default) "
                "щоб виконати. Реальний create перевірить duplicate name "
                "у parent_folder."
            ),
        }

    try:
        _check_duplicate(folder_id, name, allow_duplicate)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}

    if format == "gdoc":
        result = _create_gdoc_in_folder(
            folder_id=folder_id,
            folder_name=folder_name,
            name=name,
            initial_content=initial_content,
            project=PROJECT,
        )
    else:  # docx
        result = _create_docx_in_folder(
            folder_id=folder_id,
            folder_name=folder_name,
            name=name,
            initial_content=initial_content,
            project=PROJECT,
        )

    # _create_gdoc_in_folder вже логує через log_doc_write з tool='create_doc'.
    # _create_docx_in_folder теж. Не дублюємо лог тут.

    return result


# ----- public API: create_sheet -----

def ws_create_sheet(
    name: str,
    parent_folder: Optional[str] = None,
    format: str = "gsheet",
    columns: Optional[list[str]] = None,
    allow_duplicate: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Створює новий spreadsheet (gsheet або xlsx) у Worqen workspace.

    Args:
        name: назва нового файлу. Для xlsx суфікс .xlsx додається автоматично.
        parent_folder: name substring / Drive ID. Default — корінь Worqen Drive.
        format: 'gsheet' (default) або 'xlsx'.
        columns: опційний список header row. Якщо передано — пишеться у
            перший рядок першого аркуша.
        allow_duplicate: дозволити створення дублікату назви.
        dry_run: preview без створення.

    Returns:
        dict з ok / error / preview.
    """
    if format not in VALID_SHEET_FORMATS:
        return {
            "error": f"format має бути одним з {VALID_SHEET_FORMATS}, "
                     f"отримано: '{format}'",
            "kind": "invalid_format",
        }

    if not name or not name.strip():
        return {"error": "name не може бути порожнім", "kind": "invalid_name"}

    try:
        folder_id, folder_name = _resolve_parent_folder(parent_folder)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    if dry_run:
        return {
            "operation": "create_sheet",
            "name": name,
            "format": format,
            "parent_folder_id": folder_id,
            "parent_folder_name": folder_name,
            "columns_count": len(columns) if columns else 0,
            "columns_preview": columns[:10] if columns else None,
            "allow_duplicate": allow_duplicate,
            "would_create": True,
            "dry_run": True,
        }

    try:
        _check_duplicate(folder_id, name, allow_duplicate)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}

    if format == "gsheet":
        result = _create_gsheet_in_folder(folder_id, folder_name, name, columns)
    else:  # xlsx
        result = _create_xlsx_in_folder(folder_id, folder_name, name, columns)

    # Logging — централізовано тут (на відміну від create_doc, де gdoc/docx
    # writers логують самі).
    if result.get("ok"):
        try:
            log_doc_write(
                project=PROJECT,
                tool="ws_create_sheet",
                file_id=result["file_id"],
                file_name=result["file_name"],
                payload={
                    "format": format,
                    "folder_id": folder_id,
                    "folder_name": folder_name,
                    "columns_count": result.get("columns_count", 0),
                    "modified_after": result.get("modified_after"),
                },
            )
        except Exception as e:
            logger.warning("writes_log failed for ws_create_sheet: %s", e)

    return result


# ----- public API: create_folder -----

def ws_create_folder(
    name: str,
    parent_folder: Optional[str] = None,
    allow_duplicate: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Створює нову папку у Worqen workspace.

    Args:
        name: назва нової папки.
        parent_folder: name substring / Drive ID. Default — корінь Worqen Drive.
        allow_duplicate: дозволити створення коли у papci вже є папка з такою
            назвою.
        dry_run: preview без створення.

    Returns:
        dict з ok / error / preview.
    """
    if not name or not name.strip():
        return {"error": "name не може бути порожнім", "kind": "invalid_name"}

    try:
        folder_id, folder_name = _resolve_parent_folder(parent_folder)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    if dry_run:
        return {
            "operation": "create_folder",
            "name": name,
            "parent_folder_id": folder_id,
            "parent_folder_name": folder_name,
            "allow_duplicate": allow_duplicate,
            "would_create": True,
            "dry_run": True,
        }

    try:
        _check_duplicate(folder_id, name, allow_duplicate)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}

    service = get_service()

    try:
        new_folder = (
            service.files()
            .create(
                body={
                    "name": name,
                    "mimeType": GOOGLE_FOLDER_MIME,
                    "parents": [folder_id],
                },
                fields="id, name, mimeType, parents, webViewLink, modifiedTime",
                supportsAllDrives=True,
            )
            .execute()
        )
    except HttpError as e:
        return {"error": f"Drive create failed: {e}", "kind": "api_error"}

    new_folder_id = new_folder["id"]
    web_view_link = new_folder.get("webViewLink") or (
        f"https://drive.google.com/drive/folders/{new_folder_id}"
    )

    try:
        log_doc_write(
            project=PROJECT,
            tool="ws_create_folder",
            file_id=new_folder_id,
            file_name=name,
            payload={
                "parent_folder_id": folder_id,
                "parent_folder_name": folder_name,
                "modified_after": new_folder.get("modifiedTime"),
            },
        )
    except Exception as e:
        logger.warning("writes_log failed for ws_create_folder: %s", e)

    return {
        "ok": True,
        "folder_id": new_folder_id,
        "folder_name": name,
        "parent_folder_id": folder_id,
        "parent_folder_name": folder_name,
        "url": web_view_link,
        "modified_after": new_folder.get("modifiedTime"),
    }
