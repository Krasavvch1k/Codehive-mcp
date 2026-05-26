"""codehive_create_sheet — створення нового spreadsheet (gsheet або xlsx) у CodeHive Agency Drive.

Робить:
1. Validate name + format (gsheet/xlsx).
2. Resolve parent_folder → folder_id, folder_name (default = root).
3. Blacklist check (через writers.safety).
4. Duplicate name check (свіжий Drive query, обходить TTL кеш).
5. Маршрутизація:
   - gsheet → Drive API files.create(mimeType=spreadsheet) + опційно
     Sheets API values.update для header row.
   - xlsx → openpyxl.Workbook + Drive API multipart upload.
6. Log у writes_log.

NB: не використовуємо spreadsheets.create зі Sheets API для gsheet, бо він
кладе файл у root service account-а. Через Drive API можемо одразу вказати
parent folder.
"""

import io
import logging
from typing import Any, Optional

from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import Workbook

from projects.codehive.config import (
    CODEHIVE_ROOT_FOLDER_ID,
    GOOGLE_FOLDER_MIME,
    GOOGLE_SHEET_MIME,
    XLSX_MIME,
)
from projects.codehive.gdoc_reader import _get_folder_name, resolve_doc
from projects.codehive.writers.safety import SafetyError, check_write_allowed
from shared.drive_client import get_service as get_drive_service
from shared.sheets_client import update_values as _gsheet_update_values
from shared.writes_log import log_doc_write

logger = logging.getLogger(__name__)


VALID_SHEET_FORMATS = ("gsheet", "xlsx")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _idx_to_letter(n: int) -> str:
    """1-based column index → A1 letter."""
    if n < 1:
        raise ValueError(f"Column index must be >= 1, got {n}")
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _resolve_parent(parent_folder: Optional[str]) -> tuple[str, str]:
    """Резолвить parent_folder до (folder_id, folder_name).

    None / пусто → корінь CodeHive Agency.
    Інакше → resolve_doc(kind_filter=("folder",)).
    """
    if not parent_folder:
        if not CODEHIVE_ROOT_FOLDER_ID:
            raise ValueError("CODEHIVE_ROOT_FOLDER_ID is not set in .env.")
        root_id = CODEHIVE_ROOT_FOLDER_ID
        root_name = _get_folder_name(root_id) or "CodeHive Agency"
        return root_id, root_name

    folder_meta = resolve_doc(parent_folder, kind_filter=("folder",))
    return folder_meta["id"], folder_meta["name"]


def _check_duplicate_name(parent_id: str, name: str, mime: str) -> Optional[dict]:
    """Перевіряє чи у вказаному parent уже є файл з такою точною назвою + MIME.

    Звертається ПРЯМО до Drive API (без TTL-кешу).

    Returns:
        dict {id, name, modified} якщо знайдено дублікат, або None.
    """
    drive_service = get_drive_service()
    name_escaped = name.strip().replace("\\", "\\\\").replace("'", "\\'")
    query = (
        f"name = '{name_escaped}' "
        f"and '{parent_id}' in parents "
        f"and mimeType = '{mime}' "
        f"and trashed = false"
    )
    try:
        resp = (
            drive_service.files()
            .list(
                q=query,
                fields="files(id, name, modifiedTime)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                pageSize=10,
            )
            .execute()
        )
    except HttpError:
        return None

    files = resp.get("files", [])
    if not files:
        return None
    f = files[0]
    return {
        "id": f["id"],
        "name": f.get("name", name),
        "modified": f.get("modifiedTime"),
    }


# ---------------------------------------------------------------------------
# gsheet creation
# ---------------------------------------------------------------------------


def _create_gsheet_in_folder(
    folder_id: str,
    folder_name: str,
    name: str,
    columns: Optional[list[str]],
) -> dict[str, Any]:
    """
    Створює новий gsheet через Drive API + опційний header через Sheets API.

    Не використовуємо spreadsheets.create зі Sheets API, бо він кладе файл
    у root service account-а. Замість цього — Drive API files.create з
    mimeType=spreadsheet, а потім Sheets API values.update для header row.
    """
    drive_service = get_drive_service()

    try:
        new_file = (
            drive_service.files()
            .create(
                body={
                    "name": name,
                    "mimeType": GOOGLE_SHEET_MIME,
                    "parents": [folder_id],
                },
                fields="id, name, mimeType, parents, webViewLink, modifiedTime",
                supportsAllDrives=True,
            )
            .execute()
        )
    except HttpError as e:
        return {"error": f"Drive create failed: {e}", "kind": "api_error"}

    file_id = new_file["id"]
    web_view_link = new_file.get("webViewLink") or (
        f"https://docs.google.com/spreadsheets/d/{file_id}/edit"
    )

    columns_written = 0
    if columns:
        # Header у Sheet1!A1 — Sheets API називає default аркуш "Sheet1" /
        # локалізовано. Найбезпечніше — використати unqualified A1.
        try:
            _gsheet_update_values(
                spreadsheet_id=file_id,
                range_str=f"A1:{_idx_to_letter(len(columns))}1",
                values=[columns],
            )
            columns_written = len(columns)
        except HttpError as e:
            return {
                "ok": True,
                "warning": (
                    f"Spreadsheet створений, але header row не записався: {e}. "
                    f"Спробуй codehive_update_range з потрібним діапазоном."
                ),
                "file_id": file_id,
                "file_name": name,
                "folder_id": folder_id,
                "folder_name": folder_name,
                "url": web_view_link,
                "modified_after": new_file.get("modifiedTime"),
                "columns_count": 0,
            }

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": name,
        "folder_id": folder_id,
        "folder_name": folder_name,
        "url": web_view_link,
        "modified_after": new_file.get("modifiedTime"),
        "columns_count": columns_written,
    }


# ---------------------------------------------------------------------------
# xlsx creation
# ---------------------------------------------------------------------------


def _create_xlsx_in_folder(
    folder_id: str,
    folder_name: str,
    name: str,
    columns: Optional[list[str]],
) -> dict[str, Any]:
    """
    Створює новий xlsx у Drive folder з опційним header row.

    Будуємо файл у memory через openpyxl + multipart upload через Drive API.
    """
    wb = Workbook()
    ws = wb.active
    if columns:
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

    out = io.BytesIO()
    wb.save(out)
    content_bytes = out.getvalue()

    upload_name = name if name.lower().endswith(".xlsx") else name + ".xlsx"

    drive_service = get_drive_service()
    media = MediaIoBaseUpload(
        io.BytesIO(content_bytes),
        mimetype=XLSX_MIME,
        resumable=False,
    )

    try:
        new_file = (
            drive_service.files()
            .create(
                body={
                    "name": upload_name,
                    "mimeType": XLSX_MIME,
                    "parents": [folder_id],
                },
                media_body=media,
                fields="id, name, mimeType, parents, webViewLink, modifiedTime",
                supportsAllDrives=True,
            )
            .execute()
        )
    except HttpError as e:
        return {"error": f"Drive create failed: {e}", "kind": "api_error"}

    file_id = new_file["id"]
    web_view_link = new_file.get("webViewLink") or (
        f"https://drive.google.com/file/d/{file_id}/view"
    )

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": upload_name,
        "folder_id": folder_id,
        "folder_name": folder_name,
        "url": web_view_link,
        "modified_after": new_file.get("modifiedTime"),
        "columns_count": len(columns) if columns else 0,
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def create_sheet(
    name: str,
    parent_folder: Optional[str] = None,
    format: str = "gsheet",
    columns: Optional[list[str]] = None,
    allow_duplicate: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Створює новий spreadsheet (gsheet або xlsx) у CodeHive Agency Drive.

    Args:
        name: назва нового файлу. Для xlsx суфікс .xlsx додається автоматично.
        parent_folder: name substring / Drive ID. Default — корінь CodeHive Agency.
        format: 'gsheet' (default) або 'xlsx'.
        columns: опційний список header row. Якщо передано — пишеться у
            перший рядок першого аркуша.
        allow_duplicate: дозволити створення дублікату назви.
        dry_run: preview без створення.

    Returns:
        dict з ok / error / preview.
    """
    # 1. Validate
    if format not in VALID_SHEET_FORMATS:
        return {
            "error": (
                f"format має бути одним з {VALID_SHEET_FORMATS}, "
                f"отримано: '{format}'"
            ),
            "kind": "invalid_format",
        }

    name_clean = (name or "").strip()
    if not name_clean:
        return {"error": "name cannot be empty", "kind": "invalid_name"}

    # 2. Resolve parent
    try:
        folder_id, folder_name = _resolve_parent(parent_folder)
    except ValueError as e:
        return {"error": f"Parent folder not found: {e}", "kind": "resolve_failed"}

    # 3. Dry run preview
    if dry_run:
        return {
            "operation": "create_sheet",
            "name": name_clean,
            "format": format,
            "parent_folder_id": folder_id,
            "parent_folder_name": folder_name,
            "columns_count": len(columns) if columns else 0,
            "columns_preview": columns[:10] if columns else None,
            "allow_duplicate": allow_duplicate,
            "would_create": True,
            "dry_run": True,
        }

    # 4. Blacklist check
    try:
        check_write_allowed(
            file_id="<new>",
            file_name=name_clean,
            parent_folder_ids=[folder_id],
        )
    except SafetyError as e:
        return {"error": f"SafetyError: {e}", "kind": e.kind}

    # 5. Duplicate check
    if not allow_duplicate:
        expected_mime = GOOGLE_SHEET_MIME if format == "gsheet" else XLSX_MIME
        # Для xlsx upload_name матиме .xlsx; для duplicate check рахуємо як буде на Drive
        check_name = (
            name_clean
            if format == "gsheet" or name_clean.lower().endswith(".xlsx")
            else name_clean + ".xlsx"
        )
        dup = _check_duplicate_name(folder_id, check_name, expected_mime)
        if dup:
            return {
                "error": (
                    f"File '{check_name}' already exists in '{folder_name}'. "
                    f"Existing: id={dup['id'][:12]}..., modified={dup.get('modified')}. "
                    f"Pass allow_duplicate=true to create anyway."
                ),
                "kind": "duplicate",
                "parent_folder_id": folder_id,
                "parent_folder_name": folder_name,
                "existing_file_id": dup["id"],
            }

    # 6. Create
    if format == "gsheet":
        result = _create_gsheet_in_folder(folder_id, folder_name, name_clean, columns)
    else:
        result = _create_xlsx_in_folder(folder_id, folder_name, name_clean, columns)

    # 7. Log
    if result.get("ok"):
        try:
            log_doc_write(
                project="codehive",
                tool="create_sheet",
                file_id=result["file_id"],
                file_name=result["file_name"],
                payload={
                    "format": format,
                    "folder_id": folder_id,
                    "folder_name": folder_name,
                    "columns_count": result.get("columns_count", 0),
                    "modified_after": result.get("modified_after"),
                },
            )
        except Exception as e:
            logger.warning("writes_log failed for create_sheet: %s", e)

    return result
