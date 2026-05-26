"""
Codehive — write tools для gsheet і xlsx у CodeHive Agency.

Шар обгорток над shared.sheets_client (gsheet через Sheets API) і
shared.drive_client.upload_file_content + openpyxl (xlsx).

Pipeline:
1. Resolve query → file через gdoc_reader.resolve_doc (gsheet/xlsx only)
2. Blacklist check через codehive safety.check_write_allowed (default — порожній blacklist)
3. Drive-unchanged check (force_overwrite=True щоб обійти)
4. Маршрутизація: gsheet → sheets_client, xlsx → openpyxl + reupload
5. dry_run: повертає preview без запису
6. Logging через writes_log.log_doc_write

NB: snapshot (як у worqen) тут не робиться — CodeHive Drive не має
production-critical xlsx типу US/BUG (порожній blacklist підтверджує).
Якщо у майбутньому з'являться критичні файли — додати окремий project-level
snapshot або зареєструвати їх через WRITE_BLACKLIST_FILE_IDS.

Операції (поки):
- append_row(query, sheet, values)         — додавання в кінець таблиці

Заплановані (наступні tools):
- update_cell(query, sheet, cell, value)   — одна комірка
- update_range(query, sheet, range, values) — 2D-блок
"""

import io
import logging
import re
from copy import copy
from typing import Any, Optional

from googleapiclient.errors import HttpError
from openpyxl import load_workbook

from shared.drive_client import (
    download_file,
    fetch_current_drive_modified,
    upload_file_content,
)
from shared.safety import SafetyError, check_drive_unchanged
from shared.sheets_client import (
    append_values as _gsheet_append_values,
    update_values as _gsheet_update_values,
)
from shared.writes_log import log_doc_write

from projects.codehive.config import XLSX_MIME
from projects.codehive.gdoc_reader import resolve_doc
from projects.codehive.writers.safety import check_write_allowed

logger = logging.getLogger(__name__)


PROJECT = "codehive"
_WRITABLE_SHEET_KINDS = ("gsheet", "xlsx")


# ---------------------------------------------------------------------------
# Resolve + blacklist
# ---------------------------------------------------------------------------


def _resolve_writable_sheet(query: str) -> dict:
    """Резолвить query до gsheet/xlsx + blacklist check."""
    meta = resolve_doc(query, kind_filter=_WRITABLE_SHEET_KINDS)
    parent_id = meta.get("parent_folder_id")
    parent_ids = [parent_id] if parent_id else []
    check_write_allowed(
        file_id=meta["id"],
        file_name=meta["name"],
        parent_folder_ids=parent_ids,
    )
    return meta


# ---------------------------------------------------------------------------
# A1 helpers (shared між gsheet append/update і валідаціями)
# ---------------------------------------------------------------------------


def _validate_a1_cell(cell: str) -> None:
    if not re.match(r"^[A-Za-z]+\d+$", cell):
        raise ValueError(
            f"Невалідна A1-комірка: '{cell}'. Очікую формат 'A1', 'BC42'."
        )


def _validate_a1_range(range_str: str) -> None:
    if not re.match(r"^[A-Za-z]+\d+:[A-Za-z]+\d+$", range_str):
        raise ValueError(
            f"Невалідний A1-діапазон: '{range_str}'. Очікую формат 'A1:C10'."
        )


def _build_full_range(sheet: str, a1: str) -> str:
    """
    Складає повний range для Sheets API: 'SheetName!A1' або "'My Sheet'!A1:B10".

    Назву у лапки беремо якщо є пробіл, апостроф або початок з цифри.
    """
    needs_quote = (
        " " in sheet
        or "'" in sheet
        or (sheet and sheet[0].isdigit())
    )
    if needs_quote:
        # Екрануємо одинарні лапки всередині (Sheets вимагає подвоєння)
        escaped = sheet.replace("'", "''")
        return f"'{escaped}'!{a1}"
    return f"{sheet}!{a1}"


def _a1_cell_to_col_row(cell: str) -> tuple[int, int]:
    """A1 → (col_index_1based, row_index_1based). 'A1' → (1, 1), 'AB10' → (28, 10)."""
    cell = cell.upper()
    m = re.match(r"^([A-Z]+)(\d+)$", cell)
    if not m:
        raise ValueError(f"Невалідна A1-комірка: '{cell}'")
    letters, digits = m.group(1), m.group(2)
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return col, int(digits)


# ---------------------------------------------------------------------------
# xlsx helpers
# ---------------------------------------------------------------------------


def _open_xlsx_workbook(file_id: str) -> tuple[Any, Any]:
    """
    Завантажує xlsx з Drive (force_refresh) і відкриває у write-mode.

    Повертає (workbook, raw_bytes). Caller серіалізує і upload-ить.
    """
    raw = download_file(file_id, fmt="xlsx", force_refresh=True)
    wb = load_workbook(io.BytesIO(raw))  # default read_only=False
    return wb, raw


def _save_xlsx_and_upload(wb: Any, file_id: str) -> dict:
    """Серіалізує і заливає назад."""
    out = io.BytesIO()
    wb.save(out)
    return upload_file_content(file_id, out.getvalue(), XLSX_MIME)


def _resolve_xlsx_sheet(wb: Any, sheet_name: str) -> Any:
    """Повертає worksheet або кидає SafetyError якщо нема."""
    if sheet_name not in wb.sheetnames:
        raise SafetyError(
            f"Sheet '{sheet_name}' не знайдено. Доступні: {wb.sheetnames}",
            kind="sheet_missing",
        )
    return wb[sheet_name]


def _xlsx_find_last_data_row(ws: Any) -> int:
    """Знаходить номер останнього non-empty рядка. Якщо порожній — 0."""
    last = 0
    for row in ws.iter_rows(values_only=False):
        if all(c.value is None for c in row):
            continue
        last = row[0].row
    return last


def _xlsx_copy_row_style(ws: Any, src_row: int, dst_row: int) -> None:
    """Копіює стиль з src_row у dst_row по всіх колонках."""
    if src_row < 1:
        return
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col_idx)
        dst = ws.cell(row=dst_row, column=col_idx)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


# ---------------------------------------------------------------------------
# Public API: append_row
# ---------------------------------------------------------------------------


def append_row(
    query: str,
    sheet: str,
    values: list,
    copy_style_from_last: bool = True,
    dry_run: bool = False,
    force_overwrite: bool = False,
) -> dict[str, Any]:
    """
    Додає один рядок у кінець таблиці.

    Для gsheet — через values.append (Sheets API сам шукає кінець таблиці).
    Для xlsx — знаходимо last_data_row + 1, пишемо туди, опційно копіюємо стиль.

    Args:
        query: substring у name або Drive ID gsheet/xlsx.
        sheet: назва аркуша.
        values: 1D-list значень (не 2D). Один рядок.
        copy_style_from_last: для xlsx — копіювати стиль з попереднього рядка
            (font/fill/border/alignment). Default true.
        dry_run: preview без запису.
        force_overwrite: bypass drive-unchanged check.

    Returns:
        dict з ok / error / preview.
    """
    if not isinstance(values, list) or any(isinstance(v, list) for v in values):
        return {
            "error": "values для append_row має бути 1D-list, не 2D.",
            "kind": "invalid_values",
        }

    try:
        meta = _resolve_writable_sheet(query)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    if dry_run:
        return {
            "operation": "append_row",
            "file_id": meta["id"],
            "file_name": meta["name"],
            "kind": meta["kind"],
            "path": meta.get("path"),
            "sheet": sheet,
            "row_values": values,
            "row_length": len(values),
            "copy_style_from_last": copy_style_from_last,
            "would_write": True,
            "dry_run": True,
        }

    file_id = meta["id"]
    baseline_modified = fetch_current_drive_modified(file_id)

    if not force_overwrite:
        try:
            check_drive_unchanged(file_id, baseline_modified)
        except SafetyError as e:
            return {
                "error": str(e),
                "kind": e.kind,
                "hint": "Передай force_overwrite=True щоб ігнорувати.",
            }

    appended_row_idx: Optional[int] = None

    try:
        if meta["kind"] == "gsheet":
            # Для append range = sheet name (Sheets знайде кінець сам)
            full_range = _build_full_range(sheet, "A:A")
            api_resp = _gsheet_append_values(
                spreadsheet_id=file_id,
                range_str=full_range,
                values=[values],
            )
            modified_after = fetch_current_drive_modified(file_id)
            updates = api_resp.get("updates", {})
            result_payload = {
                "updated_range": updates.get("updatedRange"),
                "updated_cells": updates.get("updatedCells"),
            }
        else:  # xlsx
            wb, _ = _open_xlsx_workbook(file_id)
            ws = _resolve_xlsx_sheet(wb, sheet)
            last_row = _xlsx_find_last_data_row(ws)
            new_row = last_row + 1 if last_row >= 1 else 1

            if copy_style_from_last and last_row >= 1:
                _xlsx_copy_row_style(ws, last_row, new_row)

            for c_offset, val in enumerate(values, start=1):
                ws.cell(row=new_row, column=c_offset, value=val)

            upload_meta = _save_xlsx_and_upload(wb, file_id)
            modified_after = upload_meta.get("modifiedTime")
            appended_row_idx = new_row
            result_payload = {
                "row": new_row,
                "cols_written": len(values),
                "style_copied": copy_style_from_last and last_row >= 1,
            }

    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except HttpError as e:
        return {"error": f"API failed: {e}", "kind": "api_error"}
    except Exception as e:
        return {"error": f"unexpected: {e}", "kind": "unexpected"}

    try:
        log_doc_write(
            project=PROJECT,
            tool="append_row",
            file_id=file_id,
            file_name=meta["name"],
            payload={
                "kind": meta["kind"],
                "sheet": sheet,
                "row_length": len(values),
                "appended_row_idx": appended_row_idx,
                "modified_before": baseline_modified,
                "modified_after": modified_after,
                "force_overwrite": force_overwrite,
            },
        )
    except Exception as e:
        logger.warning("writes_log failed for append_row: %s", e)

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": meta["name"],
        "kind": meta["kind"],
        "sheet": sheet,
        "row_length": len(values),
        "modified_before": baseline_modified,
        "modified_after": modified_after,
        "force_overwrite_used": force_overwrite,
        **result_payload,
    }


# ---------------------------------------------------------------------------
# Public API: update_cell
# ---------------------------------------------------------------------------


def update_cell(
    query: str,
    sheet: str,
    cell: str,
    value: Any,
    dry_run: bool = False,
    force_overwrite: bool = False,
) -> dict[str, Any]:
    """
    Перезаписує ОДНУ комірку.

    Args:
        query: substring у name або повний Drive ID файла.
        sheet: назва аркуша (рядково, як у файлі).
        cell: A1-нотація однієї комірки ('A1', 'BC42').
        value: нове значення (str/int/float/bool/None).
        dry_run: preview без запису.
        force_overwrite: пропустити drive-unchanged check.

    Returns:
        dict з ok / error / preview.
    """
    try:
        _validate_a1_cell(cell)
    except ValueError as e:
        return {"error": str(e), "kind": "invalid_a1"}

    try:
        meta = _resolve_writable_sheet(query)
    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except ValueError as e:
        return {"error": str(e), "kind": "resolve_failed"}

    if dry_run:
        return {
            "operation": "update_cell",
            "file_id": meta["id"],
            "file_name": meta["name"],
            "kind": meta["kind"],
            "path": meta.get("path"),
            "sheet": sheet,
            "cell": cell.upper(),
            "value": value,
            "would_write": True,
            "dry_run": True,
        }

    file_id = meta["id"]
    baseline_modified = fetch_current_drive_modified(file_id)

    if not force_overwrite:
        try:
            check_drive_unchanged(file_id, baseline_modified)
        except SafetyError as e:
            return {
                "error": str(e),
                "kind": e.kind,
                "hint": "Передай force_overwrite=True щоб ігнорувати.",
            }

    try:
        if meta["kind"] == "gsheet":
            full_range = _build_full_range(sheet, cell.upper())
            api_resp = _gsheet_update_values(
                spreadsheet_id=file_id,
                range_str=full_range,
                values=[[value]],
            )
            modified_after = fetch_current_drive_modified(file_id)
            result_payload = {
                "updated_range": api_resp.get("updatedRange"),
                "updated_cells": api_resp.get("updatedCells"),
            }
        else:  # xlsx
            wb, _ = _open_xlsx_workbook(file_id)
            ws = _resolve_xlsx_sheet(wb, sheet)
            col, row = _a1_cell_to_col_row(cell)
            ws.cell(row=row, column=col, value=value)
            upload_meta = _save_xlsx_and_upload(wb, file_id)
            modified_after = upload_meta.get("modifiedTime")
            result_payload = {"row": row, "col": col}

    except SafetyError as e:
        return {"error": str(e), "kind": e.kind}
    except HttpError as e:
        return {"error": f"API failed: {e}", "kind": "api_error"}
    except Exception as e:
        return {"error": f"unexpected: {e}", "kind": "unexpected"}

    try:
        log_doc_write(
            project=PROJECT,
            tool="update_cell",
            file_id=file_id,
            file_name=meta["name"],
            payload={
                "kind": meta["kind"],
                "sheet": sheet,
                "cell": cell.upper(),
                "value": str(value) if value is not None else None,
                "modified_before": baseline_modified,
                "modified_after": modified_after,
                "force_overwrite": force_overwrite,
            },
        )
    except Exception as e:
        logger.warning("writes_log failed for update_cell: %s", e)

    return {
        "ok": True,
        "file_id": file_id,
        "file_name": meta["name"],
        "kind": meta["kind"],
        "sheet": sheet,
        "cell": cell.upper(),
        "value": value,
        "modified_before": baseline_modified,
        "modified_after": modified_after,
        "force_overwrite_used": force_overwrite,
        **result_payload,
    }

