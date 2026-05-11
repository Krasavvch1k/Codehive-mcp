"""
Діагностичний скрипт: показує структуру xlsx файлу.
Які листи, скільки рядків, які колонки, перші 3 рядки даних.
"""

import sys
import io
from openpyxl import load_workbook

from shared.drive_client import download_file
from projects.worqen.config import FILE_IDS


def inspect(file_key: str):
    if file_key not in FILE_IDS:
        print(f"Невідомий ключ: {file_key}")
        print(f"Доступні: {list(FILE_IDS.keys())}")
        return

    print(f"Інспектую файл: {file_key}")
    print("=" * 60)

    file_id = FILE_IDS[file_key]
    content = download_file(file_id)
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    print(f"Кількість листів: {len(wb.sheetnames)}")
    print(f"Назви листів: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"--- Лист: '{sheet_name}' ---")
        print(f"Розміри: {ws.max_row} рядків x {ws.max_column} колонок")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("(порожній)")
            print()
            continue

        headers = rows[0]
        print(f"Заголовки ({len([h for h in headers if h is not None])} непорожніх):")
        for i, h in enumerate(headers, 1):
            print(f"  {i}. {h!r}")

        print(f"\nПерші 3 рядки даних:")
        for i, row in enumerate(rows[1:4], 1):
            print(f"  Рядок {i}:")
            for col_idx, (header, value) in enumerate(zip(headers, row)):
                if value is None:
                    continue
                value_str = str(value)
                if len(value_str) > 100:
                    value_str = value_str[:97] + "..."
                print(f"    {header!r}: {value_str!r}")
            print()
        print()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Використання: python inspect_xlsx.py <file_key>")
        print(f"Доступні ключі: {[k for k in FILE_IDS if FILE_IDS[k]]}")
        sys.exit(1)

    inspect(sys.argv[1])
